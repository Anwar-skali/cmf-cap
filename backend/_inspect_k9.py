import openpyxl

path = "../CMF_K9_SCV.xlsm"
wb = openpyxl.load_workbook(path, data_only=True, read_only=False, keep_vba=True)
ws = wb["CMF K9 SCV"]
print("Sheet:", ws.title, "max_row=", ws.max_row, "max_col=", ws.max_column)
print("\n--- First 25 rows non-empty counts ---")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
    nonempty = [(j, c) for j, c in enumerate(row, start=1) if c is not None and str(c).strip() != ""]
    print(f"Row {i}: {len(nonempty)} non-empty cells | first5: {[c for _, c in nonempty[:5]]}")
print("\n--- Rows 165-173 ---")
for i, row in enumerate(ws.iter_rows(min_row=165, max_row=173, values_only=True), start=165):
    nonempty = [(j, c) for j, c in enumerate(row, start=1) if c is not None and str(c).strip() != ""]
    print(f"Row {i}: {len(nonempty)} non-empty cells | first5: {[c for _, c in nonempty[:5]]}")
print("\n--- Header row candidates: count cells identical to next row ---")
wb.close()