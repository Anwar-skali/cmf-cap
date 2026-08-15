import openpyxl

path = "../CMF_K9_SCV.xlsm"
wb = openpyxl.load_workbook(path, data_only=True, read_only=False, keep_vba=True)
print("SHEETS:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== SHEET '{name}' dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column} ===")
    if "Pilot" in name or "Suivi" in name or ws.max_row < 200:
        continue
print()
ws = wb["Pilot Sheet (Suivi)"] if "Pilot Sheet (Suivi)" in wb.sheetnames else wb[wb.sheetnames[0]]
print("Analyzing sheet:", ws.title, "max_row=", ws.max_row, "max_col=", ws.max_column)
print("\n--- First 8 rows ---")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
    nonempty = [(j, c) for j, c in enumerate(row, start=1) if c is not None and str(c).strip() != ""]
    print(f"Row {i}: {len(nonempty)} non-empty | {nonempty[:6]}")
print("\n--- Last 10 rows (check trailing rows) ---")
for i, row in enumerate(ws.iter_rows(min_row=ws.max_row - 9, max_row=ws.max_row, values_only=True), start=ws.max_row - 9):
    nonempty = [(j, c) for j, c in enumerate(row, start=1) if c is not None and str(c).strip() != ""]
    print(f"Row {i}: {len(nonempty)} non-empty | {nonempty[:5]}")
wb.close()
