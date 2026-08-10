import io
import logging
import re
import time
from typing import Any
import openpyxl

logger = logging.getLogger(__name__)

HEADER_KEYWORDS = {
    "unique_id", "unique id", "supplier_name", "supplier name", "part_number", "part number",
    "part_name", "part name", "apqp", "weekly_capacity", "weekly capacity", "sqe", "sqm",
    "cat1", "cat2", "cat3", "capacity", "supplier", "project", "code", "status", "title",
    "name", "description", "target", "unit", "quantity", "date", "type", "category",
    "email", "phone", "address", "version", "id"
}

# Sheet name or content patterns that strongly suggest non-data sheets (dashboards, summaries, charts, pivot tables)
PENALTY_PATTERNS = re.compile(
    r"\b(dashboard|chart|charts|graph|pivot|pivot table|pivot tables|summary|overview|cover|index|"
    r"kpi|kpi summary|kpi carry-over|report|template|readme|guide|instruction|legend|info|"
    r"master|lookup|reference|ref|drop|dropdown|list|validation|config|sqd list|run scheduling|gst|grand total|row labels|status summary)\b",
    re.IGNORECASE,
)


class ExcelHeaderExtractor:
    """
    Intelligent Excel Extractor with:
    1. Worksheet scoring — automatically finds the data sheet (not dashboards/charts/summaries/pivots)
    2. Header row detection — scores top 20 rows within selected sheet
    3. User override for both worksheet and header row
    4. Zero row data leakage to LLMs
    """

    # ─── Row scoring ──────────────────────────────────────────────────────────

    @staticmethod
    def _score_row(row_cells: list[Any], template_keywords: set[str] | None = None) -> tuple[float, float]:
        """
        Scores a single row. Returns tuple of (raw_score, confidence_percentage).
        """
        non_empty = [str(c).strip() for c in row_cells if c is not None and str(c).strip() != ""]
        if not non_empty:
            return 0.0, 0.0

        score = float(len(non_empty))
        all_keywords = HEADER_KEYWORDS | (template_keywords or set())

        numeric_count = 0
        keyword_hits = 0
        for val in non_empty:
            clean_v = val.lower().replace("_", " ").strip()
            if clean_v in all_keywords:
                score += 10.0
                keyword_hits += 1
            else:
                for kw in all_keywords:
                    if len(kw) >= 3 and (kw in clean_v or clean_v in kw):
                        score += 4.0
                        keyword_hits += 1
                        break

            if re.match(r"^\d+(\.\d+)?$", val) or re.match(r"^\d{4}-\d{2}-\d{2}", val):
                numeric_count += 1

        if len(non_empty) > 0 and (numeric_count / len(non_empty)) > 0.4:
            score -= 20.0

        raw_score = max(score, 0.0)
        # Confidence calculation: ratio of keyword hits & text length
        conf_pct = min(100.0, max(10.0, round((keyword_hits / max(len(non_empty), 1)) * 70.0 + (len(non_empty) / 30.0) * 30.0, 1))) if raw_score > 0 else 0.0
        return raw_score, conf_pct

    # ─── Orientation Detection ───────────────────────────────────────────────

    @staticmethod
    def detect_orientation(
        rows_sample: list[list[Any]],
        template_keywords: set[str] | None = None,
        specified_orientation: str | None = None,
    ) -> dict[str, Any]:
        """
        Detects whether worksheet layout is VERTICAL (Col A = Field, Col B = Value)
        or HORIZONTAL (Row N = Headers, Rows N+1.. = Records).
        """
        if specified_orientation and str(specified_orientation).lower() in ("vertical", "horizontal"):
            orient = str(specified_orientation).upper()
            return {
                "orientation": orient,
                "orientation_confidence": 100.0,
                "field_column": "A" if orient == "VERTICAL" else "",
                "value_column": "B" if orient == "VERTICAL" else "",
                "reason": f"User-specified {orient.capitalize()}.",
            }

        all_keywords = HEADER_KEYWORDS | (template_keywords or set())

        col_a_matches = 0
        col_b_populated = 0
        total_rows_checked = 0

        for r in rows_sample:
            if not r:
                continue
            cell_a = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
            cell_b = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""

            if cell_a:
                total_rows_checked += 1
                clean_a = cell_a.lower().replace("_", " ").strip()
                if clean_a in all_keywords or any(len(kw) >= 3 and kw in clean_a for kw in all_keywords):
                    col_a_matches += 1
                if cell_b:
                    col_b_populated += 1

        max_row_headers = 0
        for r in rows_sample[:5]:
            non_empty = [str(c).strip() for c in r if c is not None and str(c).strip() != ""]
            matches = 0
            for val in non_empty:
                clean = val.lower().replace("_", " ").strip()
                if clean in all_keywords or any(len(kw) >= 3 and kw in clean for kw in all_keywords):
                    matches += 1
            if len(non_empty) >= 5 and matches >= 3:
                max_row_headers = max(max_row_headers, matches)

        is_vertical = False
        confidence = 50.0
        reason = ""

        if col_a_matches >= 4 and col_b_populated >= 4 and col_a_matches > max_row_headers:
            is_vertical = True
            ratio = col_a_matches / max(total_rows_checked, 1)
            confidence = min(98.0, max(70.0, round(ratio * 100.0, 1)))
            reason = f"{col_a_matches} recognized project fields in Column A, {col_b_populated} values in Column B."
        else:
            confidence = 85.0 if max_row_headers >= 3 else 60.0
            reason = "Standard horizontal table row layout detected."

        orientation = "VERTICAL" if is_vertical else "HORIZONTAL"

        logger.info(
            "[ORIENTATION DETECTED] %s (Conf: %.1f%%) | Field Col: %s | Value Col: %s | Reason: %s",
            orientation, confidence, "A" if is_vertical else "", "B" if is_vertical else "", reason
        )

        return {
            "orientation": orientation,
            "orientation_confidence": confidence,
            "field_column": "A" if is_vertical else "",
            "value_column": "B" if is_vertical else "",
            "reason": reason,
        }

    # ─── Sheet classification & scoring ─────────────────────────────────────

    @staticmethod
    def _classify_worksheet(
        sheet_name: str,
        populated_rows: int,
        max_cols: int,
        pivot_indicators: int,
        project_field_matches: int,
        is_dashboard_name: bool,
    ) -> str:
        if populated_rows < 2:
            return "EMPTY"
        if is_dashboard_name:
            clean_name = sheet_name.lower()
            if "kpi" in clean_name:
                return "KPI"
            if "pivot" in clean_name:
                return "PIVOT_TABLE"
            if "lookup" in clean_name or "ref" in clean_name or "guide" in clean_name:
                return "REFERENCE_DATA"
            return "SUMMARY"
        if pivot_indicators >= 2:
            return "PIVOT_TABLE"
        if max_cols <= 5 and pivot_indicators >= 1:
            return "SUMMARY"
        if max_cols <= 5 and project_field_matches < 3:
            return "KPI"
        if project_field_matches >= 3 or max_cols >= 10:
            return "PROJECT_DATA"
        return "UNKNOWN"

    @staticmethod
    def _score_sheet_vertical(
        rows_sample: list[list[Any]],
        sheet_name: str,
        template_keywords: set[str] | None = None,
        is_dashboard_name: bool = False,
        orientation_info: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """
        Scores a worksheet for VERTICAL (key-value) orientation:
        Col A = field names, Col B = values.
        Scored by populated field/value pairs, domain matches in Col A, fill ratio, and empty value ratio.
        """
        name_penalty = -60.0 if is_dashboard_name else 0.0
        all_keywords = HEADER_KEYWORDS | (template_keywords or set())

        field_pairs = 0
        field_matches = 0
        value_fills = 0
        duplicate_fields = 0
        seen_fields: set[str] = set()
        preview_cells: list[str] = []

        for row in rows_sample:
            cell_a = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            cell_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if not cell_a:
                continue

            field_pairs += 1
            if cell_b:
                value_fills += 1

            clean_a = cell_a.lower().replace("_", " ").strip()
            if clean_a in seen_fields:
                duplicate_fields += 1
            else:
                seen_fields.add(clean_a)

            if clean_a in all_keywords or any(len(kw) >= 3 and (kw in clean_a or clean_a in kw) for kw in all_keywords):
                field_matches += 1

            if len(preview_cells) < 4:
                preview_cells.append(f"{cell_a}: {cell_b}" if cell_b else cell_a)

        fill_ratio = value_fills / max(field_pairs, 1)
        match_ratio = field_matches / max(field_pairs, 1)

        raw_score = field_matches * 15.0 + field_pairs * 2.0 + fill_ratio * 40.0 - duplicate_fields * 15.0 + name_penalty
        classification = "PROJECT_DATA" if field_matches >= 3 else ("UNKNOWN" if field_pairs >= 3 else "EMPTY")

        confidence_pct = min(99.0, max(5.0, round(match_ratio * 60.0 + fill_ratio * 30.0 + min(field_pairs / 20.0, 1.0) * 10.0, 1))) if field_pairs >= 3 else 5.0

        orient_info = orientation_info or {
            "orientation": "VERTICAL",
            "orientation_confidence": confidence_pct,
            "field_column": "A",
            "value_column": "B",
            "reason": f"{field_matches} recognized project fields in Column A, {value_fills} values in Column B.",
        }

        return {
            "sheet_name": sheet_name,
            "score": round(max(raw_score, 0.0), 2),
            "confidence": confidence_pct,
            "populated_rows": field_pairs,
            "max_columns": 2,
            "keyword_hits": field_matches,
            "is_dashboard_name": is_dashboard_name,
            "classification": classification,
            "pivot_indicators": 0,
            "project_field_matches": field_matches,
            "structure_similarity": round(fill_ratio * 100.0, 1),
            "orientation": orient_info["orientation"],
            "orientation_confidence": orient_info["orientation_confidence"],
            "field_column": orient_info["field_column"],
            "value_column": orient_info["value_column"],
            "orientation_reason": orient_info["reason"],
            "preview": preview_cells,
        }

    @staticmethod
    def _score_sheet(
        sheet: Any,
        sheet_name: str,
        template_keywords: set[str] | None = None,
        specified_orientation: str | None = None,
    ) -> dict[str, Any]:
        """
        Scores a worksheet on how likely it contains tabular project data.
        Returns a dict with classification, score, confidence %, and metadata.
        """
        is_dashboard_name = bool(PENALTY_PATTERNS.search(sheet_name))
        name_penalty = -60.0 if is_dashboard_name else 0.0

        rows_sample: list[list[Any]] = []
        try:
            for r in sheet.iter_rows(max_row=20, values_only=True):
                rows_sample.append(list(r) if r else [])
        except Exception:
            pass

        orientation_info = ExcelHeaderExtractor.detect_orientation(rows_sample, template_keywords, specified_orientation)

        if orientation_info["orientation"] == "VERTICAL":
            return ExcelHeaderExtractor._score_sheet_vertical(
                rows_sample, sheet_name, template_keywords, is_dashboard_name, orientation_info
            )

        if not rows_sample:
            return {
                "sheet_name": sheet_name,
                "score": 0.0,
                "confidence": 5.0,
                "populated_rows": 0,
                "max_columns": 0,
                "keyword_hits": 0,
                "is_dashboard_name": is_dashboard_name,
                "classification": "EMPTY",
                "pivot_indicators": 0,
                "project_field_matches": 0,
                "structure_similarity": 0.0,
                "orientation": orientation_info["orientation"],
                "orientation_confidence": orientation_info["orientation_confidence"],
                "field_column": orientation_info["field_column"],
                "value_column": orientation_info["value_column"],
                "orientation_reason": orientation_info["reason"],
                "preview": [],
            }

        non_empty_rows = 0
        max_cols = 0
        keyword_hits = 0
        pivot_indicators = 0
        project_field_matches = 0
        all_keywords = HEADER_KEYWORDS | (template_keywords or set())

        pivot_content_regex = re.compile(
            r"\b(grand total|row labels|count of|nombre de|total|kpi|sum of|average of|count|kpi summary)\b",
            re.IGNORECASE,
        )

        for row in rows_sample:
            non_empty_cells = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
            if non_empty_cells:
                non_empty_rows += 1
                max_cols = max(max_cols, len(non_empty_cells))
                for cell_val in non_empty_cells:
                    clean = cell_val.lower().replace("_", " ").strip()
                    if pivot_content_regex.search(clean):
                        pivot_indicators += 1
                    if clean in all_keywords:
                        keyword_hits += 1
                        project_field_matches += 1
                    else:
                        for kw in all_keywords:
                            if len(kw) >= 3 and (kw in clean or clean in kw):
                                keyword_hits += 1
                                project_field_matches += 1
                                break

        classification = ExcelHeaderExtractor._classify_worksheet(
            sheet_name, non_empty_rows, max_cols, pivot_indicators, project_field_matches, is_dashboard_name
        )

        col_score = min(max_cols, 80) * 4.0
        row_score = min(non_empty_rows, 20) * 1.5
        kw_density = min((keyword_hits / max(max_cols, 1)) * 40.0, 100.0)

        structure_similarity = 0.0
        if template_keywords and len(template_keywords) > 0:
            match_count = sum(1 for kw in template_keywords if any(kw in str(row).lower() for row in rows_sample[:5]))
            structure_similarity = min(1.0, match_count / max(len(template_keywords), 1))

        struct_bonus = structure_similarity * 60.0

        classification_penalty = 0.0
        if classification == "PIVOT_TABLE":
            classification_penalty = -100.0
        elif classification == "KPI":
            classification_penalty = -120.0
        elif classification == "SUMMARY":
            classification_penalty = -80.0
        elif classification == "REFERENCE_DATA":
            classification_penalty = -40.0

        pivot_penalty = min(pivot_indicators, 10) * -15.0
        narrow_penalty = -25.0 if max_cols < 6 and orientation_info["orientation"] != "VERTICAL" else 0.0

        raw_score = col_score + row_score + kw_density + struct_bonus + name_penalty + classification_penalty + pivot_penalty + narrow_penalty

        max_benchmark = 320.0 + 30.0 + 100.0 + 60.0  # 510.0
        if raw_score <= 0:
            confidence_pct = 5.0 if is_dashboard_name or classification != "PROJECT_DATA" else 10.0
        else:
            confidence_pct = min(99.0, max(5.0, round((raw_score / max_benchmark) * 100.0, 1)))

        preview_cells: list[str] = []
        for row in rows_sample[:10]:
            cells = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
            if len(cells) > len(preview_cells):
                preview_cells = cells[:8]

        logger.info(
            "[SHEET ANALYSIS] Sheet: '%s' | Class: %s | Orient: %s | Cols: %d | Rows: %d | Field Matches: %d | Score: %.1f | Conf: %.1f%%",
            sheet_name, classification, orientation_info["orientation"], max_cols, non_empty_rows, project_field_matches, raw_score, confidence_pct
        )

        return {
            "sheet_name": sheet_name,
            "score": round(raw_score, 2),
            "confidence": confidence_pct,
            "populated_rows": non_empty_rows,
            "max_columns": max_cols,
            "keyword_hits": keyword_hits,
            "is_dashboard_name": is_dashboard_name,
            "classification": classification,
            "pivot_indicators": pivot_indicators,
            "project_field_matches": project_field_matches,
            "structure_similarity": round(structure_similarity * 100, 1),
            "orientation": orientation_info["orientation"],
            "orientation_confidence": orientation_info["orientation_confidence"],
            "field_column": orientation_info["field_column"],
            "value_column": orientation_info["value_column"],
            "orientation_reason": orientation_info["reason"],
            "preview": preview_cells,
        }

    # ─── Public API ───────────────────────────────────────────────────────────

    @staticmethod
    def extract_workbook_info(
        file_bytes: bytes,
        template_keywords: set[str] | None = None,
        specified_orientation: str | None = None,
    ) -> dict[str, Any]:
        """
        Reads workbook and returns worksheet scores WITHOUT extracting headers.
        Used in Step 2 of the UI to let user pick the correct worksheet.
        """
        start_t = time.perf_counter()
        if not file_bytes:
            raise ValueError("Empty file payload received.")

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        sheet_names = workbook.sheetnames
        if not sheet_names:
            workbook.close()
            raise ValueError("The uploaded Excel workbook contains no worksheets.")

        sheet_scores: list[dict[str, Any]] = []
        best_sheet = sheet_names[0]
        best_score = -9999.0

        for name in sheet_names:
            try:
                ws = workbook[name]
                info = ExcelHeaderExtractor._score_sheet(ws, name, template_keywords, specified_orientation)
            except Exception as e:
                logger.debug("[SHEET SCORE] Failed scoring sheet '%s': %s", name, e)
                info = {
                    "sheet_name": name, "score": 0.0, "confidence": 5.0,
                    "populated_rows": 0, "max_columns": 0, "keyword_hits": 0,
                    "is_dashboard_name": False, "preview": [],
                    "orientation": "HORIZONTAL", "orientation_confidence": 50.0,
                    "field_column": "", "value_column": "", "orientation_reason": "",
                }
            sheet_scores.append(info)
            if info["score"] > best_score:
                best_score = info["score"]
                best_sheet = name

        workbook.close()
        duration_ms = round((time.perf_counter() - start_t) * 1000, 2)

        logger.info(
            "[WORKBOOK SCAN] %d sheets scanned in %.2f ms. Best sheet: '%s' (score=%.1f)",
            len(sheet_names), duration_ms, best_sheet, best_score,
        )
        return {
            "sheets": sheet_scores,
            "detected_sheet": best_sheet,
            "duration_ms": duration_ms,
        }

    @staticmethod
    def extract_headers_with_details(
        file_bytes: bytes,
        specified_header_row: int | None = None,
        specified_sheet_name: str | None = None,
        template_keywords: set[str] | None = None,
        specified_orientation: str | None = None,
    ) -> tuple[list[str], int, float, str, float, list[dict[str, Any]], float, dict[str, Any]]:
        """
        Reads a specific (or auto-detected) worksheet, scores top 20 rows, and extracts headers.

        Returns:
          - headers: list[str]
          - detected_row_index: int (1-based)
          - header_confidence: float (0-100%)
          - sheet_used: str
          - sheet_confidence: float (0-100%)
          - row_previews: list[dict] (preview of top 20 rows)
          - duration_ms: float
          - orientation_info: dict (orientation, confidence, field_column, value_column, reason)
        """
        start_t = time.perf_counter()
        if not file_bytes:
            raise ValueError("Empty file payload received.")

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

        sheet_names = workbook.sheetnames
        if not sheet_names:
            workbook.close()
            raise ValueError("The uploaded Excel workbook contains no worksheets.")

        sheet_info_map = {}
        best_name = sheet_names[0]
        best_sheet_score = -9999.0

        for name in sheet_names:
            try:
                ws = workbook[name]
                info = ExcelHeaderExtractor._score_sheet(ws, name, template_keywords, specified_orientation)
                sheet_info_map[name] = info
                if info["score"] > best_sheet_score:
                    best_sheet_score = info["score"]
                    best_name = name
            except Exception:
                pass

        if specified_sheet_name and specified_sheet_name in sheet_names:
            target_sheet_name = specified_sheet_name
        else:
            target_sheet_name = best_name

        target_info = sheet_info_map.get(target_sheet_name, {})
        sheet_confidence = target_info.get("confidence", 85.0)

        sheet = workbook[target_sheet_name]
        top_20_rows: list[list[Any]] = []

        try:
            for r in sheet.iter_rows(max_row=50, values_only=True):
                top_20_rows.append(list(r) if r else [])
        except Exception as e:
            logger.debug("Failed iterating rows on sheet '%s': %s", target_sheet_name, e)

        workbook.close()

        orientation_info = ExcelHeaderExtractor.detect_orientation(top_20_rows, template_keywords, specified_orientation)

        if not top_20_rows:
            return [], 1, 0.0, target_sheet_name, sheet_confidence, [], round((time.perf_counter() - start_t) * 1000, 2), orientation_info

        headers: list[str] = []
        detected_row_idx = 1
        header_confidence = 90.0
        row_previews: list[dict[str, Any]] = []

        if orientation_info["orientation"] == "VERTICAL":
            # Vertical Format: Col A = Field Name, Col B = Value
            seen_fields = set()
            duplicate_fields = []

            for idx, r in enumerate(top_20_rows, start=1):
                if not r:
                    continue
                cell_a = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
                cell_b = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""

                if cell_a:
                    if cell_a in seen_fields:
                        duplicate_fields.append((cell_a, idx))
                    else:
                        seen_fields.add(cell_a)
                    headers.append(cell_a)

                if idx <= 20:
                    row_previews.append({
                        "row_number": idx,
                        "score": 10.0 if cell_a else 0.0,
                        "confidence": 95.0 if cell_a else 0.0,
                        "non_empty_count": 2 if (cell_a and cell_b) else (1 if cell_a or cell_b else 0),
                        "preview": [cell_a, cell_b][:2],
                    })

            if duplicate_fields:
                dup_str = ", ".join([f"'{field}' (Row {r_num})" for field, r_num in duplicate_fields[:3]])
                logger.warning("[VERTICAL EXCEL] Duplicate field names detected in Column A: %s", dup_str)

            header_confidence = min(99.0, max(75.0, round((len(headers) / max(len(top_20_rows), 1)) * 100.0, 1)))

        else:
            # Horizontal Format: Row N = Header Row
            best_row_idx = 1
            best_row_score = -999.0
            best_row_conf = 0.0

            for idx, row in enumerate(top_20_rows[:20], start=1):
                score, conf = ExcelHeaderExtractor._score_row(row, template_keywords)
                non_empty_cells = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
                row_previews.append({
                    "row_number": idx,
                    "score": round(score, 2),
                    "confidence": conf,
                    "non_empty_count": len(non_empty_cells),
                    "preview": non_empty_cells[:8],
                })
                if score > best_row_score:
                    best_row_score = score
                    best_row_idx = idx
                    best_row_conf = conf

            target_row_idx = (
                specified_header_row
                if specified_header_row and 1 <= specified_header_row <= len(top_20_rows)
                else best_row_idx
            )
            detected_row_idx = target_row_idx
            target_row = top_20_rows[target_row_idx - 1]
            header_confidence = (
                row_previews[target_row_idx - 1]["confidence"]
                if 1 <= target_row_idx <= len(row_previews)
                else best_row_conf
            )

            for cell in target_row:
                val = str(cell).strip() if cell is not None else ""
                if val or headers:
                    headers.append(val)

            while headers and headers[-1] == "":
                headers.pop()

        duration_ms = round((time.perf_counter() - start_t) * 1000, 2)
        logger.info(
            "[HEADER DETECTED] Sheet: '%s' | Orient: %s | Headers: %d | Elapsed: %.2f ms",
            target_sheet_name, orientation_info["orientation"], len(headers), duration_ms,
        )
        return headers, detected_row_idx, header_confidence, target_sheet_name, sheet_confidence, row_previews, duration_ms, orientation_info

    @staticmethod
    def extract_headers_with_timing(file_bytes: bytes) -> tuple[list[str], float]:
        headers, _, _, _, _, _, duration_ms, _ = ExcelHeaderExtractor.extract_headers_with_details(file_bytes)
        return headers, duration_ms

    @staticmethod
    def extract_headers_from_bytes(file_bytes: bytes) -> list[str]:
        headers, _ = ExcelHeaderExtractor.extract_headers_with_timing(file_bytes)
        return headers

