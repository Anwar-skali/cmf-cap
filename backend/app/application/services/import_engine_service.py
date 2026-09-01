from __future__ import annotations

import io
import time
import uuid
from datetime import datetime, date
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.domain.enums import ActivityAction
from app.domain.import_schema import ENTITY_IMPORT_SCHEMAS, EntityImportSchema, ImportColumnSpec
from app.application.services.data_normalizer import DataNormalizer, NormalizationResult
from app.application.services.header_normalizer import normalize_header
from app.infrastructure.persistence.unit_of_work import UnitOfWork
import logging

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# K0 Make Battery – authoritative field definitions
# ---------------------------------------------------------------------------
# All 42 real internal K0 schema fields in logical order (Buyer → Capacity → SQD).
# Columns K, L, M, N (#REF!) and O (libre) in the raw Excel are unmapped dead columns.
K0_FIELD_KEYS: list[str] = [
    "part_number", "index", "description", "coef", "serial_piece_price",
    "mass_purchase", "ru", "noa", "make_battery_lp_1", "make_battery_lp_2",
    "supplier_name", "vendor_cofor", "manufacturer_cofor", "combined_cofor",
    "tango_order", "ei_status", "comments", "week_project_target_1",
    "forecast_week_1", "completed_week_1", "week_project_target_2",
    "forecast_week_2", "completed_week_2", "week_project_target_3",
    "forecast_week_3", "completed_week_3", "quality", "supply_chain",
    "global_purchasing", "cpl", "rcpi", "minimum_quality_status_acted",
    "mass_inquired", "packaging_readiness_unlweb_validated",
    "tango_contract_validated", "supplier_capability_confirmed",
    "it_cpl_corail_setting", "fcla_validates", "ple_created", "edi_opened",
    "um_logistic_flow_validated", "manufacturing_process_validated",
]

# ---------------------------------------------------------------------------
# The 47 SOURCE column positions present in the real K0 Excel file (Row 8, Cols A–AU).
#
# The K0 source workbook ("Pilot Sheet (Suivi)") has 47 physical columns (A through AU).
# - Columns A..J (0..9) map to Buyer identification fields.
# - Columns K..O (10..14) are dead/empty columns (#REF! and libre) -> None.
# - Columns P..V (15..21) map to Buyer supplier/status/comments fields.
# - Columns W..AE (22..30) map to Capacity Manager milestones 1, 2, 3.
# - Columns AF..AU (31..46) map to SQD assessments & validations.
# ---------------------------------------------------------------------------
K0_SOURCE_COLUMNS: list[str | None] = [
    "part_number",        # Col A   (idx=0)  – Part Number
    "index",              # Col B   (idx=1)  – Index
    "description",        # Col C   (idx=2)  – Description
    "coef",               # Col D   (idx=3)  – Coef.
    "serial_piece_price", # Col E   (idx=4)  – Serial Piece Price
    "mass_purchase",      # Col F   (idx=5)  – Mass Purchase
    "ru",                 # Col G   (idx=6)  – RU
    "noa",                # Col H   (idx=7)  – NOA
    "make_battery_lp_1",  # Col I   (idx=8)  – Make Battery (LP) — 1st occurrence
    "make_battery_lp_2",  # Col J   (idx=9)  – Make Battery (LP) — 2nd occurrence
    None,                 # Col K   (idx=10) – #REF! (dead column)
    None,                 # Col L   (idx=11) – #REF! (dead column)
    None,                 # Col M   (idx=12) – #REF! (dead column)
    None,                 # Col N   (idx=13) – #REF! (dead column)
    None,                 # Col O   (idx=14) – libre (dead column)
    "supplier_name",      # Col P   (idx=15) – Supplier name
    "vendor_cofor",       # Col Q   (idx=16) – Vendor COFOR
    "manufacturer_cofor", # Col R   (idx=17) – Manufacturer COFOR
    "combined_cofor",     # Col S   (idx=18) – Combined COFOR
    "tango_order",        # Col T   (idx=19) – Tango order
    "ei_status",          # Col U   (idx=20) – EI status
    "comments",           # Col V   (idx=21) – Comments
    "week_project_target_1",  # Col W  (idx=22) – Week\nProject Target (1st)
    "forecast_week_1",        # Col X  (idx=23) – Forecast\nWeek (1st)
    "completed_week_1",       # Col Y  (idx=24) – Completed\nWeek (1st)
    "week_project_target_2",  # Col Z  (idx=25) – Week\nProject Target (2nd)
    "forecast_week_2",        # Col AA (idx=26) – Forecast\nWeek (2nd)
    "completed_week_2",       # Col AB (idx=27) – Completed\nWeek (2nd)
    "week_project_target_3",  # Col AC (idx=28) – Week\nProject Target (3rd)
    "forecast_week_3",        # Col AD (idx=29) – Forecast\nWeek (3rd)
    "completed_week_3",       # Col AE (idx=30) – Completed\nWeek (3rd)
    "quality",            # Col AF  (idx=31) – Quality
    "supply_chain",       # Col AG  (idx=32) – Supply Chain
    "global_purchasing",  # Col AH  (idx=33) – Global Purchasing
    "cpl",                # Col AI  (idx=34) – CPL
    "rcpi",               # Col AJ  (idx=35) – RCPI
    "minimum_quality_status_acted",       # Col AK (idx=36) – Minimum Quality status acted
    "mass_inquired",                      # Col AL (idx=37) – Mass inquired
    "packaging_readiness_unlweb_validated", # Col AM (idx=38) – Packaging readiness, UNLweb validated
    "tango_contract_validated",           # Col AN (idx=39) – TANGO contract validated
    "supplier_capability_confirmed",      # Col AO (idx=40) – Supplier capability confirmed
    "it_cpl_corail_setting",              # Col AP (idx=41) – IT CPL (CORAIL,) setting
    "fcla_validates",                     # Col AQ (idx=42) – FCLA Validates
    "ple_created",                        # Col AR (idx=43) – PLE created
    "edi_opened",                         # Col AS (idx=44) – EDI opened
    "um_logistic_flow_validated",         # Col AT (idx=45) – UM logistic flow validated
    "manufacturing_process_validated",      # Col AU (idx=46) – Manufacturing process validated
]
K0_SOURCE_COLUMN_COUNT: int = len(K0_SOURCE_COLUMNS)  # == 47
K0_ALL_FIELD_KEYS = K0_FIELD_KEYS  # alias for backward compat

# K0 Excel sheet and header row settings
K0_SHEET_NAME = "Pilot Sheet (Suivi)"
K0_HEADER_ROW = 8         # 1-indexed Excel row containing English headers
K0_DATA_START_ROW = 9     # 1-indexed Excel row of first real data record
K0_UNIQUE_KEY_COL = 0     # 0-indexed column of part_number (Column A)

# Known K0 Excel header names (from Row 8) mapped deterministically to field keys.
# NOTE: This dict is used ONLY for display/metadata purposes now.
# The actual row conversion uses K0_SOURCE_COLUMNS with index-based lookup,
# which is immune to duplicate-header dict collision bugs.
K0_HEADER_TO_KEY: dict[str, str] = {
    "Part Number": "part_number",
    "Index": "index",
    "Description": "description",
    "Coef.": "coef",
    "Coef": "coef",
    "Serial Piece Price": "serial_piece_price",
    "Serial Piece Price ": "serial_piece_price",
    "Mass Purchase\n(Coef. x Price per part)": "mass_purchase",
    "Mass Purchase": "mass_purchase",
    "RU": "ru",
    "NOA": "noa",
    "Make Battery (LP)": "make_battery_lp_1",
    "Make Battery (LP) 1": "make_battery_lp_1",
    "Make Battery (LP) 2": "make_battery_lp_2",
    "Supplier Name": "supplier_name",
    "Supplier name": "supplier_name",
    "Vendor COFOR": "vendor_cofor",
    "Vendor Cofor": "vendor_cofor",
    "Manufacturer COFOR": "manufacturer_cofor",
    "Manufacturer Cofor": "manufacturer_cofor",
    "Combined COFOR": "combined_cofor",
    "Tango Order": "tango_order",
    "EI Status": "ei_status",
    "EI status": "ei_status",
    "Comments": "comments",
    "Week\nProject Target": "week_project_target_1",
    "Week Project Target 1": "week_project_target_1",
    "Forecast\nWeek": "forecast_week_1",
    "Forecast Week 1": "forecast_week_1",
    "Completed\nWeek": "completed_week_1",
    "Completed Week 1": "completed_week_1",
    "Week Project Target 2": "week_project_target_2",
    "Forecast Week 2": "forecast_week_2",
    "Completed Week 2": "completed_week_2",
    "Week Project Target 3": "week_project_target_3",
    "Forecast Week 3": "forecast_week_3",
    "Completed Week 3": "completed_week_3",
    "Quality": "quality",
    "Supply Chain": "supply_chain",
    "Global Purchasing": "global_purchasing",
    "CPL": "cpl",
    "RCPI": "rcpi",
    "Minimum Quality status acted": "minimum_quality_status_acted",
    "Minimum Quality Status Acted": "minimum_quality_status_acted",
    "Mass inquired": "mass_inquired",
    "Mass Inquired": "mass_inquired",
    "Packaging readiness, UNLweb \nvalidated": "packaging_readiness_unlweb_validated",
    "Packaging Readiness UNLWEB Validated": "packaging_readiness_unlweb_validated",
    "TANGO contract validated": "tango_contract_validated",
    "Tango Contract Validated": "tango_contract_validated",
    "Supplier capability confirmed": "supplier_capability_confirmed",
    "Supplier Capability Confirmed": "supplier_capability_confirmed",
    "IT CPL (CORAIL,) setting": "it_cpl_corail_setting",
    "IT CPL (CORAIL) Setting": "it_cpl_corail_setting",
    "IT CPL Corail Setting": "it_cpl_corail_setting",
    "FCLA validates": "fcla_validates",
    "FCLA Validates": "fcla_validates",
    "PLE created": "ple_created",
    "PLE Created": "ple_created",
    "EDI opened": "edi_opened",
    "EDI Opened": "edi_opened",
    "UM logistic flow validated": "um_logistic_flow_validated",
    "UM Logistic Flow Validated": "um_logistic_flow_validated",
    "Manufacturing process validated": "manufacturing_process_validated",
    "Manufacturing Process Validated": "manufacturing_process_validated",
}

# Core structure codes that have their own fast-path positional mapping
_K0_CODES = {"K0", "K0_MAKE_BATTERY"}


def _k0_col_letter(idx: int) -> str:
    """Return the Excel column letter(s) for a 0-based column index (A, B, …, Z, AA, AB, …)."""
    if idx < 26:
        return chr(ord("A") + idx)
    # Two-letter columns: A=0 → AA=26, AB=27, …, AZ=51, BA=52, …
    first = (idx - 26) // 26
    second = (idx - 26) % 26
    return chr(ord("A") + first) + chr(ord("A") + second)


def _build_k0_index_mapping(source_headers: list[str]) -> dict[int, str]:
    """
    Build a deterministic **index-based** mapping for K0 source columns.

    Returns a dict mapping 0-based column index -> K0 internal field key,
    covering ONLY the first K0_SOURCE_COLUMN_COUNT (47) columns.

    WHY INDEX-BASED (not header-string-based):
    The K0 source Excel has duplicate header strings:
    - Columns W, Z, AC (idx 22, 25, 28) all share header "Week\nProject Target"
    - Columns X, AA, AD (idx 23, 26, 29) all share header "Forecast\nWeek"
    - Columns Y, AB, AE (idx 24, 27, 30) all share header "Completed\nWeek"
    - Columns I & J (idx 8 & 9) both share header "Make Battery (LP)"
    A dict keyed on header-string would be overwritten on each duplicate.
    Index-based mapping is completely immune to this class of bug.

    SCOPE: The first 47 columns (A–AU) are source data columns.
    Columns 48+ are empty Excel cells and are never imported.

    This runs in O(1) for each column access and never calls Ollama/AI.
    """
    mapping: dict[int, str] = {}
    col_count = min(len(source_headers), K0_SOURCE_COLUMN_COUNT)
    for col_idx in range(col_count):
        mapping[col_idx] = K0_SOURCE_COLUMNS[col_idx]

    # Emit structured debug log showing the full column mapping
    logger.info(
        "[K0 PARSE] header_row=%d source_columns=%d (mapped=%d from file)",
        K0_HEADER_ROW, K0_SOURCE_COLUMN_COUNT, col_count,
    )
    for col_idx, field_key in mapping.items():
        hdr = source_headers[col_idx] if col_idx < len(source_headers) else "(no header)"
        col_letter = _k0_col_letter(col_idx)
        logger.debug("[K0 MAPPING] Col %s (idx=%d) / Header %r -> field '%s'",
                     col_letter, col_idx, hdr, field_key)
    return mapping


def _build_k0_column_mapping(headers: list[str]) -> dict[str, str | None]:
    """
    LEGACY shim kept for backward compatibility with callers that pass
    ``column_mapping.get(header)`` using a string key.

    Delegates to _build_k0_index_mapping and wraps the result as a
    header-string dict so existing call sites continue to work.
    NOTE: Duplicate headers (e.g. "Week\nProject Target" on cols W, Z, AC)
    will collide in a string-keyed dict — the LAST occurrence wins.  That is
    acceptable for the string-keyed shim because the actual row conversion
    loop uses k0_index_map (index-based) which is collision-free.
    """
    index_map = _build_k0_index_mapping(headers)
    result: dict[str, str | None] = {}
    for col_idx, field_key in index_map.items():
        if col_idx < len(headers):
            result[headers[col_idx]] = field_key
    return result


class ImportEngineService:
    def __init__(self, uow: UnitOfWork) -> None:
        self._uow = uow

    @staticmethod
    def _make_json_safe(record: dict[str, Any]) -> dict[str, Any]:
        """Deep-convert a record dict to JSON-serializable primitives.

        SQLAlchemy stores the 'data' column as JSON. Python types that
        are not natively JSON-serializable (datetime.date, datetime.datetime,
        Decimal) must be converted to strings before being passed to the ORM.
        """
        result: dict[str, Any] = {}
        for k, v in record.items():
            if isinstance(v, datetime):
                result[k] = v.isoformat()
            elif isinstance(v, date):
                result[k] = v.isoformat()
            elif isinstance(v, Decimal):
                result[k] = float(v)
            elif isinstance(v, dict):
                result[k] = ImportEngineService._make_json_safe(v)
            elif isinstance(v, list):
                result[k] = [
                    ImportEngineService._make_json_safe(i) if isinstance(i, dict)
                    else (i.isoformat() if isinstance(i, (date, datetime)) else i)
                    for i in v
                ]
            else:
                result[k] = v
        return result

    def parse_excel_file(
        self,
        file_bytes: bytes,
        specified_orientation: str | None = None,
        specified_sheet_name: str | None = None,
        specified_header_row: int | None = None,
        unique_key_col_index: int | None = None,
        max_source_columns: int | None = None,
    ) -> dict[str, Any]:
        """
        Extract sheet names, headers, and raw rows from an uploaded Excel file.
        Supports both HORIZONTAL (tabular) and VERTICAL (key-value) orientations.

        Args:
            specified_header_row: 1-indexed row number of the header row.  When provided
                the scanner skips straight to that row instead of auto-detecting.
            unique_key_col_index: 0-indexed column position of the primary key column
                (e.g. 0 for ``part_number`` in K0).  When supplied, any data row whose
                unique-key cell is empty / zero / formula-only is treated as an empty
                row and excluded from the returned ``rows`` list.
            max_source_columns: When set, headers and row values are truncated to this
                many columns after the header row is determined.  Use this for K0 to
                restrict ingestion to the 26 authoritative source columns (A–Z) and
                prevent Excel helper/formula columns (AA+) from being read as data.
        """
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheets = workbook.sheetnames
        if not sheets:
            raise ValueError("The Excel file contains no worksheets.")

        best_sheet_name = specified_sheet_name if specified_sheet_name and specified_sheet_name in sheets else sheets[0]
        best_rows: list[tuple[Any, ...]] = []

        if specified_sheet_name and specified_sheet_name in sheets:
            ws = workbook[specified_sheet_name]
            best_rows = list(ws.iter_rows(values_only=True))
        else:
            max_count = 0
            for sheet_name in sheets:
                ws = workbook[sheet_name]
                sheet_rows = list(ws.iter_rows(values_only=True))
                if not sheet_rows:
                    continue

                for r in sheet_rows[:20]:
                    non_empty_count = len([c for c in r if c is not None and str(c).strip() != ""])
                    if non_empty_count > max_count:
                        max_count = non_empty_count
                        best_sheet_name = sheet_name
                        best_rows = sheet_rows

        if not best_rows:
            return {
                "sheets": sheets,
                "headers": [],
                "rows": [],
                "total_rows": 0,
                "physical_rows": 0,
                "empty_rows_count": 0,
                "orientation": "HORIZONTAL",
            }

        # Run orientation detection
        from app.application.services.excel_header_extractor import ExcelHeaderExtractor
        orient_info = ExcelHeaderExtractor.detect_orientation(
            [list(r) for r in best_rows[:50]], specified_orientation=specified_orientation
        )
        orientation = orient_info["orientation"]

        if orientation == "VERTICAL":
            headers: list[str] = []
            single_row_vals: list[Any] = []
            seen_fields = set()

            for r in best_rows:
                if not r:
                    continue
                cell_a = str(r[0]).strip() if len(r) > 0 and r[0] is not None else ""
                cell_b = r[1] if len(r) > 1 else None

                if cell_a:
                    if cell_a in seen_fields:
                        logger.warning("[VERTICAL PARSER] Duplicate field '%s' detected in Column A.", cell_a)
                    else:
                        seen_fields.add(cell_a)
                    headers.append(cell_a)
                    single_row_vals.append(cell_b)

            return {
                "sheets": sheets,
                "selected_sheet": best_sheet_name,
                "headers": headers,
                "rows": [single_row_vals] if headers else [],
                "total_rows": 1 if headers else 0,
                "physical_rows": 1 if headers else 0,
                "empty_rows_count": 0,
                "orientation": "VERTICAL",
                "orientation_info": orient_info,
            }

        # HORIZONTAL Mode
        # Determine header row index (0-based into best_rows)
        if specified_header_row and 1 <= specified_header_row <= len(best_rows):
            header_idx = specified_header_row - 1
            logger.info("[PARSE] Using specified_header_row=%d (0-based idx=%d)", specified_header_row, header_idx)
        else:
            # Auto-detect: row with the most non-empty cells in first 20 rows
            header_idx = 0
            max_count = 0
            for idx, r in enumerate(best_rows[:20]):
                non_empty_count = len([c for c in r if c is not None and str(c).strip() != ""])
                if non_empty_count > max_count:
                    max_count = non_empty_count
                    header_idx = idx
            logger.info("[PARSE] Auto-detected header row %d (0-based idx=%d)", header_idx + 1, header_idx)

        header_row = best_rows[header_idx]
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        while headers and headers[-1] == "":
            headers.pop()

        # K0 source-column boundary: the K0 workbook has exactly 47 real source
        # columns (A–AU).  Columns beyond AU are empty Excel cells and must never
        # be treated as importable data.  max_source_columns enforces this boundary.
        if max_source_columns is not None and len(headers) > max_source_columns:
            logger.info(
                "[K0 PARSE] Capping headers from %d to %d real source columns",
                len(headers), max_source_columns,
            )
            headers = headers[:max_source_columns]

        # Sentinel values produced by Excel formulas on blank rows that look non-empty
        _FORMULA_SENTINELS: set[str] = {"0", "0.0", "#ref!", "#n/a", "#value!", "#name?", "none", "nan"}

        def _cell_is_meaningful(val: Any) -> bool:
            """Return True if the cell carries a real human-entered value."""
            if val is None:
                return False
            s = str(val).strip()
            return s != "" and s.lower() not in _FORMULA_SENTINELS

        def _row_is_empty(row_vals: list[Any], uk_col_idx: int | None) -> bool:
            """
            A row is EMPTY when:
            - All cells are None / "" / formula sentinels, OR
            - A unique-key column index is given and that cell is not meaningful
              (even if other formula cells look non-empty).
            """
            if uk_col_idx is not None and 0 <= uk_col_idx < len(row_vals):
                return not _cell_is_meaningful(row_vals[uk_col_idx])
            return not any(_cell_is_meaningful(c) for c in row_vals)

        data_rows: list[list[Any]] = []
        # We track the furthest offset that had at least one raw non-None cell
        # (to measure the physical used range of the sheet).
        last_physical_offset = -1
        for offset, row in enumerate(best_rows[header_idx + 1:]):
            row_vals = list(row[: len(headers)])  # already bounded by truncated headers
            if len(row_vals) < len(headers):
                row_vals.extend([None] * (len(headers) - len(row_vals)))

            # Check whether the physical row has ANY content (incl. formulas)
            has_any_content = any(c is not None and str(c).strip() != "" for c in row_vals)
            if has_any_content:
                last_physical_offset = offset

            if _row_is_empty(row_vals, unique_key_col_index):
                # Bug fix: only count as 'empty' if the row is within the physical used range.
                # Rows completely beyond last_physical_offset are trailing nulls, not empty records.
                if has_any_content:
                    physical_row_num = header_idx + 2 + offset  # 1-based Excel row
                    logger.debug(
                        "[IMPORT ROW FILTER] physical_row=%d status=EMPTY_IGNORED",
                        physical_row_num,
                    )
            else:
                physical_row_num = header_idx + 2 + offset
                logger.debug(
                    "[IMPORT ROW FILTER] physical_row=%d status=DATA",
                    physical_row_num,
                )
                data_rows.append(row_vals)

        physical_rows = last_physical_offset + 1
        # empty_rows_count = rows within the physical range that had formula/content
        # but no valid unique key (i.e. were not real data rows).
        empty_rows_count = physical_rows - len(data_rows)

        logger.info(
            "[IMPORT ROW FILTER] sheet=%s header_row=%d physical_rows=%d data_rows=%d empty_rows=%d",
            best_sheet_name, header_idx + 1, physical_rows, len(data_rows), empty_rows_count,
        )

        return {
            "sheets": sheets,
            "selected_sheet": best_sheet_name,
            "headers": headers,
            "rows": data_rows,
            "total_rows": len(data_rows),
            "physical_rows": physical_rows,
            "empty_rows_count": empty_rows_count,
            "orientation": "HORIZONTAL",
            "orientation_info": orient_info,
        }

    def auto_map_columns(
        self, headers: list[str], schema: EntityImportSchema
    ) -> dict[str, str | None]:
        """
        Deterministic multi-level auto-matching between Excel headers and schema fields:
        Level 1: Exact normalized match (key / label)
        Level 2: Known aliases match
        Level 3: Fuzzy matching (>= 0.90 confidence)
        """
        from app.application.services.header_normalizer import normalize_header, compute_similarity, strip_module_prefix

        mapping: dict[str, str | None] = {}
        assigned_keys: set[str] = set()

        # Pass 1: Exact matches
        for header in headers:
            norm_header = normalize_header(header)
            if not norm_header:
                mapping[header] = None
                continue

            stripped_header = strip_module_prefix(norm_header)
            matched_key: str | None = None
            for col in schema.columns:
                if col.key in assigned_keys:
                    continue
                norm_col_key = normalize_header(col.key)
                norm_col_label = normalize_header(col.label)
                if (
                    norm_header == norm_col_key
                    or norm_header == norm_col_label
                    or (stripped_header and (stripped_header == norm_col_key or stripped_header == norm_col_label))
                ):
                    matched_key = col.key
                    assigned_keys.add(col.key)
                    break

            if matched_key:
                mapping[header] = matched_key

        # Pass 2: Known alias matches
        for header in headers:
            if header in mapping:
                continue
            norm_header = normalize_header(header)
            if not norm_header:
                mapping[header] = None
                continue

            stripped_header = strip_module_prefix(norm_header)
            matched_key = None
            for col in schema.columns:
                if col.key in assigned_keys:
                    continue
                norm_aliases = [normalize_header(a) for a in col.aliases if a]
                if norm_header in norm_aliases or (stripped_header and stripped_header in norm_aliases):
                    matched_key = col.key
                    assigned_keys.add(col.key)
                    break

            if matched_key:
                mapping[header] = matched_key

        # Pass 3: High confidence fuzzy matches (>= 0.90)
        for header in headers:
            if header in mapping:
                continue
            norm_header = normalize_header(header)
            if not norm_header:
                mapping[header] = None
                continue

            best_key = None
            best_score = 0.0
            for col in schema.columns:
                if col.key in assigned_keys:
                    continue
                score_k = compute_similarity(norm_header, normalize_header(col.key))
                score_l = compute_similarity(norm_header, normalize_header(col.label))
                score_a = max([compute_similarity(norm_header, normalize_header(a)) for a in col.aliases if a], default=0.0)
                max_score = max(score_k, score_l, score_a)
                if max_score > best_score:
                    best_score = max_score
                    best_key = col.key

            if best_key and best_score >= 0.90:
                mapping[header] = best_key
                assigned_keys.add(best_key)
            else:
                mapping[header] = None

        return mapping

    @staticmethod
    def _normalize_column_mapping(
        custom_mapping: dict[str, Any], schema: EntityImportSchema, headers: list[str]
    ) -> dict[str, str]:
        """
        Normalizes any custom_mapping dict (whether Excel->DB, DB->Excel, or Ollama dict)
        into a consistent {excel_header: db_field_key} mapping dict.
        """
        from app.application.services.header_normalizer import normalize_header

        column_mapping: dict[str, str] = {}
        schema_keys = {col.key for col in schema.columns}
        schema_key_map = {col.key.lower(): col.key for col in schema.columns}
        schema_norm_map = {normalize_header(col.key): col.key for col in schema.columns}

        # Build lookup for Excel headers present in the uploaded file
        header_map = {normalize_header(h): h for h in headers if h}

        for k, v in custom_mapping.items():
            if v == "__ignore__" or k == "__ignore__":
                continue

            if isinstance(v, dict):
                excel_col = v.get("excel")
                if excel_col and excel_col != "__ignore__":
                    column_mapping[str(excel_col)] = k
            elif isinstance(v, str):
                clean_k = str(k).strip()
                clean_v = str(v).strip()

                norm_k = normalize_header(clean_k)
                norm_v = normalize_header(clean_v)

                k_is_header = norm_k in header_map
                v_is_header = norm_v in header_map

                exact_k_header = header_map.get(norm_k, clean_k) if k_is_header else clean_k
                exact_v_header = header_map.get(norm_v, clean_v) if v_is_header else clean_v

                k_is_db = clean_k in schema_keys or norm_k in schema_norm_map or clean_k.lower() in schema_key_map
                v_is_db = clean_v in schema_keys or norm_v in schema_norm_map or clean_v.lower() in schema_key_map

                resolved_k_db = schema_norm_map.get(norm_k, schema_key_map.get(clean_k.lower(), clean_k)) if k_is_db else clean_k
                resolved_v_db = schema_norm_map.get(norm_v, schema_key_map.get(clean_v.lower(), clean_v)) if v_is_db else clean_v

                if k_is_header and v_is_db:
                    column_mapping[exact_k_header] = resolved_v_db
                elif v_is_header and k_is_db:
                    column_mapping[exact_v_header] = resolved_k_db
                elif k_is_header:
                    column_mapping[exact_k_header] = resolved_v_db
                elif v_is_header:
                    column_mapping[exact_v_header] = resolved_k_db
                else:
                    column_mapping[clean_k] = resolved_v_db if v_is_db else clean_v

        return column_mapping

    async def _get_schema(self, entity_type: str) -> EntityImportSchema:
        lower_type = entity_type.lower().strip()
        if lower_type in ENTITY_IMPORT_SCHEMAS:
            return ENTITY_IMPORT_SCHEMAS[lower_type]

        # Resolve dynamic template (K9, K0, or custom template)
        from app.application.services.template_context_service import TemplateContextService

        ctx_svc = TemplateContextService(self._uow)
        ctx = await ctx_svc.get_template_context(entity_type)
        if not ctx.fields:
            raise ValueError(f"Unsupported entity or template type: '{entity_type}'")

        columns = []
        for f in ctx.fields:
            lower_k = f.key.lower()
            is_unique = (
                lower_k in [
                    "unique_id", "code", "project_code", "part_number", "part_no",
                    "project_id", "ref", "project_ref", "pn", "id"
                ]
                or "code" in lower_k
                or "unique" in lower_k
                or "part_number" in lower_k
            )
            columns.append(
                ImportColumnSpec(
                    key=f.key,
                    label=f.label,
                    required=f.required,
                    type=f.type if f.type in ["string", "integer", "number", "date", "week", "enum", "textarea", "text"] else "string",
                    aliases=f.aliases,
                    description=f.description,
                    unique_key=is_unique,
                )
            )

        return EntityImportSchema(
            entity_type=ctx.template_code,
            display_name=ctx.template_name,
            columns=columns,
            sample_rows=[],
        )

    @staticmethod
    def _unique_key_columns(schema: EntityImportSchema) -> list[str]:
        """Return the schema keys that identify a real record (unique keys).

        ``id`` is excluded: it is never supplied by an Excel import, so treating
        it as a record-identifying key would wrongly classify rows as records.
        """
        keys = [c.key for c in schema.columns if c.unique_key and c.key != "id"]
        if not keys:
            for c in schema.columns:
                lower_k = c.key.lower()
                if (
                    lower_k in (
                        "project_code", "code", "unique_id", "part_number", "part_no",
                        "project_id", "ref", "project_ref", "pn"
                    )
                    or "code" in lower_k
                    or "unique" in lower_k
                ):
                    keys.append(c.key)
        return keys

    @staticmethod
    def _extract_project_code(record: dict[str, Any], schema: EntityImportSchema | None = None) -> str | None:
        """Extracts the unique project code/identifier from a mapped record dictionary.

        Priority order:
        1. Schema's declared unique_key column (template-specific, e.g. unique_id for K9, part_number for K0)
        2. Known generic unique-identifier keys as fallback
        3. Any key containing 'code', 'unique', or 'part_number' as last resort
        """
        # 1. Schema unique_key takes highest priority — ensures each template uses its own declared key
        if schema:
            for col in schema.columns:
                if col.unique_key and col.key != "id":
                    val = record.get(col.key)
                    if val is not None and str(val).strip() != "":
                        return str(val).strip()

        # 2. Generic fallback keys (only reached when no schema unique_key found)
        for k in (
            "unique_id",
            "code",
            "part_number",
            "part_no",
            "project_id",
            "ref",
            "project_ref",
            "pn",
            "id",
        ):
            val = record.get(k)
            if val is not None and str(val).strip() != "":
                return str(val).strip()

        # 3. Last resort: any key containing 'code', 'unique', or 'part_number'
        for k, val in record.items():
            if val is not None and str(val).strip() != "":
                lower_k = k.lower()
                if "unique" in lower_k or "part_number" in lower_k:
                    return str(val).strip()

        return None


    @staticmethod
    def _extract_project_name(record: dict[str, Any], code: str | None = None) -> str:
        """Extracts the project name/title from a mapped record dictionary."""
        for k in (
            "project_name",
            "name",
            "part_name",
            "title",
            "project_title",
            "label",
        ):
            val = record.get(k)
            if val is not None and str(val).strip() != "":
                return str(val).strip()

        for k, val in record.items():
            if val is not None and str(val).strip() != "":
                lower_k = k.lower()
                if "name" in lower_k or "title" in lower_k:
                    return str(val).strip()

        return f"Project {code}" if code else "Untitled Project"

    @staticmethod
    def _extract_project_description(record: dict[str, Any]) -> str | None:
        """Extracts project description/info/notes/comments from a mapped record dictionary."""
        for k in (
            "description",
            "project_description",
            "part_info",
            "notes",
            "comments",
            "comment",
            "details",
        ):
            val = record.get(k)
            if val is not None and str(val).strip() != "":
                return str(val).strip()
        return None

    @staticmethod
    def _extract_client_name(record: dict[str, Any]) -> str | None:
        """Extracts customer/client name from a mapped record dictionary."""
        for k in ("customer", "client_name", "client", "customer_name", "company"):
            val = record.get(k)
            if val is not None and str(val).strip() != "":
                return str(val).strip()
        return None

    @staticmethod
    def _row_has_record_key(values: dict[str, Any], schema: EntityImportSchema) -> bool:
        """Determine whether a parsed row represents an actual data record.

        A row is a record when any unique-key column carries a non-empty value
        (e.g. ``part_number`` for the K0 structure). Rows without a unique-key
        value are non-record rows (section headers, notes, blank placeholders)
        and must be excluded from validation, preview, and import counts.
        """
        unique_key_keys = ImportEngineService._unique_key_columns(schema)
        if not unique_key_keys:
            return any(
                v is not None and str(v).strip() != ""
                for v in values.values()
            )
        return any(
            values.get(k) is not None and str(values.get(k)).strip() != ""
            for k in unique_key_keys
        )

    async def preview_and_validate(
        self,
        file_bytes: bytes,
        file_name: str,
        entity_type: str,
        custom_mapping: dict[str, str] | None = None,
        orientation: str | None = None,
        sheet_name: str | None = None,
        header_row: int | None = None,
    ) -> dict[str, Any]:
        """
        Parses the Excel file, validates header schema and row values,
        checks for duplicates, and produces a complete validation preview report.
        """
        is_k0 = entity_type.upper() in _K0_CODES
        start_perf = time.perf_counter()

        # K0 fast-path: use known sheet/header/key settings
        effective_header_row = header_row
        effective_sheet = sheet_name
        effective_uk_col: int | None = None
        if is_k0:
            effective_header_row = effective_header_row or K0_HEADER_ROW
            effective_sheet = effective_sheet or K0_SHEET_NAME
            effective_uk_col = K0_UNIQUE_KEY_COL

        schema = await self._get_schema(entity_type)
        parsed = self.parse_excel_file(
            file_bytes,
            specified_orientation=orientation,
            specified_sheet_name=effective_sheet,
            specified_header_row=effective_header_row,
            unique_key_col_index=effective_uk_col,
            max_source_columns=K0_SOURCE_COLUMN_COUNT if is_k0 else None,
        )
        headers = parsed["headers"]
        data_rows = parsed["rows"]
        parsed_orientation = parsed.get("orientation", "HORIZONTAL")
        parsed_orientation_info = parsed.get("orientation_info", {})

        # Step 1: Mapping setup
        if is_k0:
            k0_index_map = _build_k0_index_mapping(headers)
            if custom_mapping:
                column_mapping = self._normalize_column_mapping(custom_mapping, schema, headers)
            else:
                column_mapping = _build_k0_column_mapping(headers)

            elapsed_ms = round((time.perf_counter() - start_perf) * 1000, 2)
            mapped_count = len([v for v in k0_index_map.values() if v])
            unmapped_count = K0_SOURCE_COLUMN_COUNT - mapped_count
            logger.info(
                "[K0 PARSE]\nsheet=%s\nheader_row=%d\nsource_columns=%d",
                effective_sheet or K0_SHEET_NAME, K0_HEADER_ROW, K0_SOURCE_COLUMN_COUNT,
            )
            logger.info(
                "[K0 DATA]\nphysical_rows=%d\ndata_rows=%d\nempty_rows=%d",
                parsed.get("physical_rows", 0), len(data_rows), parsed.get("empty_rows_count", 0),
            )
            logger.info(
                "[K0 MAPPING]\nmapped_columns=%d\nunmapped_columns=%d",
                mapped_count, unmapped_count,
            )
        elif custom_mapping:
            column_mapping = self._normalize_column_mapping(custom_mapping, schema, headers)
        else:
            column_mapping = self.auto_map_columns(headers, schema)

        # Invert mapping to map db_field_key -> excel_header
        if is_k0:
            mapped_fields: dict[str, str] = {
                field_key: headers[col_idx] if col_idx < len(headers) else field_key
                for col_idx, field_key in k0_index_map.items() if field_key
            }
        else:
            mapped_fields = {
                v: k for k, v in column_mapping.items() if v is not None
            }

        validation_errors: list[dict[str, Any]] = []

        # Step 2: Header checks
        missing_mandatory_fields: list[str] = []
        for col_spec in schema.columns:
            if col_spec.required and col_spec.key not in mapped_fields:
                missing_mandatory_fields.append(col_spec.label)

        if missing_mandatory_fields:
            validation_errors.append(
                {
                    "row_index": 0,
                    "column_name": "Header",
                    "field_key": "header",
                    "raw_value": None,
                    "error_type": "MissingMandatoryColumn",
                    "message": f"Missing required column(s): {', '.join(missing_mandatory_fields)}",
                }
            )

        # Step 3: Row-by-Row validation, duplicate detection, and action classification
        unique_key_spec = next((c for c in schema.columns if c.unique_key), None)
        seen_unique_values: dict[str, int] = {}  # clean_value -> first_seen_row_idx

        # Query existing database records (including soft-deleted)
        existing_db_records: dict[str, dict[str, Any]] = {}
        if unique_key_spec:
            existing_db_records = await self._get_existing_db_records(entity_type)

        is_project_entity = entity_type.lower() not in ("suppliers", "parts", "risks", "capacity")

        valid_rows_count = 0
        error_rows_set: set[int] = set()
        duplicate_in_excel_set: set[int] = set()
        duplicate_in_db_set: set[int] = set()
        new_rows_set: set[int] = set()
        update_rows_set: set[int] = set()
        restore_rows_set: set[int] = set()
        validation_error_rows_set: set[int] = set()
        non_record_rows_set: set[int] = set()
        record_rows_count = 0

        preview_rows: list[dict[str, Any]] = []
        normalization_warnings: list[dict[str, Any]] = []

        for row_idx, row_vals in enumerate(data_rows, start=1):
            row_raw_dict: dict[str, Any] = {}
            row_dict: dict[str, Any] = {}
            row_norm_details: dict[str, Any] = {}
            row_has_error = False

            # Convert row using column mapping.
            # For K0: use index-based mapping (k0_index_map) so duplicate header strings
            # (e.g. "Week\nProject Target" on cols W, Z, AC) can never overwrite each other.
            # Rows are already bounded to K0_SOURCE_COLUMN_COUNT by parse_excel_file.
            # For all other templates: use the standard header-string mapping.
            if is_k0:
                for col_idx, val in enumerate(row_vals):
                    db_key = k0_index_map.get(col_idx)
                    if db_key:
                        row_raw_dict[db_key] = val
            else:
                for header, val in zip(headers, row_vals):
                    db_key = column_mapping.get(header)
                    if not db_key:
                        norm_h = normalize_header(header)
                        for m_hdr, m_k in column_mapping.items():
                            if m_hdr and normalize_header(m_hdr) == norm_h:
                                db_key = m_k
                                break
                    if db_key:
                        row_raw_dict[db_key] = val

            is_record = self._row_has_record_key(row_raw_dict, schema)
            if is_record:
                record_rows_count += 1
            else:
                non_record_rows_set.add(row_idx)

            # Normalize and Validate each schema column
            for col_spec in schema.columns:
                raw_val = row_raw_dict.get(col_spec.key)
                excel_hdr = mapped_fields.get(col_spec.key, col_spec.label)

                # Stage 1: Data Normalization
                norm = DataNormalizer.normalize(raw_val, col_spec)

                if norm.warning:
                    normalization_warnings.append(
                        {
                            "row_index": row_idx,
                            "column_name": col_spec.label,
                            "field_key": col_spec.key,
                            "original_value": str(raw_val) if raw_val is not None else None,
                            "warning": norm.warning,
                        }
                    )

                # Stage 2: Validation on Normalized Value
                if norm.is_null:
                    if col_spec.required:
                        msg = f"Field '{col_spec.label}' is mandatory."
                        validation_errors.append(
                            {
                                "row_index": row_idx,
                                "column_name": col_spec.label,
                                "field_key": col_spec.key,
                                "raw_value": str(raw_val) if raw_val is not None else None,
                                "error_type": "EmptyValue",
                                "message": msg,
                            }
                        )
                        row_has_error = True
                        validation_error_rows_set.add(row_idx)
                        row_norm_details[col_spec.key] = {
                            "original": raw_val,
                            "normalized": None,
                            "status": "error",
                            "message": msg,
                        }
                    else:
                        row_dict[col_spec.key] = None
                        row_norm_details[col_spec.key] = {
                            "original": raw_val,
                            "normalized": None,
                            "status": "warning" if norm.warning else "valid",
                            "message": norm.warning or "Normalized to NULL",
                        }
                else:
                    coerced_val, err_msg = self._validate_and_coerce_type(norm.normalized_value, col_spec)
                    if err_msg:
                        validation_errors.append(
                            {
                                "row_index": row_idx,
                                "column_name": col_spec.label,
                                "field_key": col_spec.key,
                                "raw_value": str(raw_val),
                                "error_type": "InvalidDataType",
                                "message": err_msg,
                            }
                        )
                        row_has_error = True
                        validation_error_rows_set.add(row_idx)
                        row_norm_details[col_spec.key] = {
                            "original": raw_val,
                            "normalized": str(norm.normalized_value),
                            "status": "error",
                            "message": err_msg,
                        }
                    else:
                        row_dict[col_spec.key] = coerced_val
                        norm_disp = coerced_val.isoformat() if isinstance(coerced_val, (date, datetime)) else coerced_val
                        row_norm_details[col_spec.key] = {
                            "original": raw_val,
                            "normalized": norm_disp,
                            "status": "valid",
                            "message": "OK",
                        }

                logger.debug(
                    "[NORM] Row %d | Excel Header: '%s' -> DB Field: '%s' (%s) | Original: %r | Normalized: %r | Result: %s | Reason: %s",
                    row_idx,
                    excel_hdr,
                    col_spec.key,
                    col_spec.type,
                    raw_val,
                    row_dict.get(col_spec.key),
                    "ERROR" if row_norm_details[col_spec.key]["status"] == "error" else "OK",
                    row_norm_details[col_spec.key]["message"],
                )

            # Stage 3: Duplicate detection and Action Classification (CREATE / UPDATE / RESTORE)
            row_action = "CREATE"
            row_status = "new"

            if unique_key_spec and unique_key_spec.key in row_dict and row_dict[unique_key_spec.key] is not None:
                u_val = str(row_dict[unique_key_spec.key]).strip()
                clean_u_val = u_val.lower()

                # Duplicate inside same Excel file check
                if clean_u_val in seen_unique_values:
                    first_row = seen_unique_values[clean_u_val]
                    validation_errors.append(
                        {
                            "row_index": row_idx,
                            "column_name": unique_key_spec.label,
                            "field_key": unique_key_spec.key,
                            "raw_value": u_val,
                            "error_type": "DuplicateInFile",
                            "message": f"Duplicate value '{u_val}' found in Excel (first seen in row {first_row}).",
                        }
                    )
                    row_has_error = True
                    duplicate_in_excel_set.add(row_idx)
                else:
                    seen_unique_values[clean_u_val] = row_idx

                # Check Database existence (Active vs Soft-Deleted)
                if clean_u_val in existing_db_records:
                    rec_info = existing_db_records[clean_u_val]
                    is_soft_del = bool(rec_info.get("is_deleted", False))
                    duplicate_in_db_set.add(row_idx)
                    row_dict["_exists_in_db"] = True
                    row_dict["_is_deleted"] = is_soft_del

                    if is_soft_del:
                        row_action = "RESTORE"
                        row_status = "deleted"
                        restore_rows_set.add(row_idx)
                    else:
                        row_action = "UPDATE"
                        row_status = "existing"
                        update_rows_set.add(row_idx)

                    # For non-project entities in strict insert mode, record DuplicateInDatabase
                    if not is_project_entity:
                        validation_errors.append(
                            {
                                "row_index": row_idx,
                                "column_name": unique_key_spec.label,
                                "field_key": unique_key_spec.key,
                                "raw_value": u_val,
                                "error_type": "DuplicateInDatabase",
                                "message": f"Value '{u_val}' already exists in database.",
                            }
                        )
                        row_has_error = True
                        validation_error_rows_set.add(row_idx)
                else:
                    row_action = "CREATE"
                    row_status = "new"
                    new_rows_set.add(row_idx)

            if is_record:
                if row_has_error:
                    error_rows_set.add(row_idx)
                else:
                    valid_rows_count += 1

            if len(preview_rows) < 100:
                preview_entry: dict[str, Any] = {}
                for k, v in row_dict.items():
                    preview_entry[k] = v.isoformat() if isinstance(v, (date, datetime)) else v
                preview_entry["_row_index"] = row_idx
                preview_entry["_has_error"] = row_has_error
                preview_entry["_is_record"] = is_record
                preview_entry["_action"] = row_action
                preview_entry["_status"] = row_status
                preview_entry["norm_details"] = row_norm_details
                preview_entry["_norm_details"] = row_norm_details
                preview_rows.append(preview_entry)

        # Non-record rows are ignored entirely
        if non_record_rows_set:
            validation_errors = [
                e for e in validation_errors if e.get("row_index") not in non_record_rows_set
            ]
            normalization_warnings = [
                w for w in normalization_warnings if w.get("row_index") not in non_record_rows_set
            ]

        valid_records = set(range(1, len(data_rows) + 1)) - error_rows_set - non_record_rows_set

        if is_k0:
            logger.info("[K0 PREVIEW]\nstatus=success")

        return {
            "entity_type": entity_type,
            "entity_display_name": schema.display_name,
            "file_name": file_name,
            "total_rows": len(data_rows),
            "physical_rows": parsed.get("physical_rows", len(data_rows)),
            "empty_rows_count": parsed.get("empty_rows_count", 0),
            "non_record_rows_count": len(non_record_rows_set),
            "record_rows_count": record_rows_count,
            "valid_rows_count": valid_rows_count,
            "error_rows_count": len(error_rows_set),
            "new_count": len(new_rows_set & valid_records),
            "update_count": len(update_rows_set & valid_records),
            "restore_count": len(restore_rows_set & valid_records),
            "duplicate_in_excel_count": len(duplicate_in_excel_set),
            "duplicate_in_db_count": len(duplicate_in_db_set),
            "validation_errors_count": len(
                {e.get("row_index") for e in validation_errors if e.get("row_index", 0) > 0}
            ),
            "headers": headers,
            "column_mapping": column_mapping,
            "available_schema_columns": [
                {
                    "key": c.key,
                    "label": c.label,
                    "required": c.required,
                    "type": c.type,
                }
                for c in schema.columns
            ],
            "validation_errors": validation_errors,
            "normalization_warnings": normalization_warnings,
            "preview_rows": preview_rows,
            "orientation": parsed_orientation,
            "orientation_info": parsed_orientation_info,
        }

    async def execute_import(
        self,
        file_bytes: bytes,
        file_name: str,
        entity_type: str,
        column_mapping: dict[str, str],
        mode: str = "insert",  # insert | upsert
        strategy: str = "skip_invalid",  # skip_invalid | rollback_all
        user_id: uuid.UUID | None = None,
        user_email: str | None = None,
        orientation: str | None = None,
        sheet_name: str | None = None,
        header_row: int | None = None,
    ) -> dict[str, Any]:
        """
        Executes database import with transactional safety, audit logging,
        and history recording.
        """
        is_k0 = entity_type.upper() in _K0_CODES
        start_time = time.time()
        logger.info("[IMPORT DEBUG] Starting execution for file '%s' with template '%s'", file_name, entity_type)
        logger.info("[IMPORT DEBUG] Selected template: '%s'", entity_type)
        logger.info("[IMPORT DEBUG] Uploaded filename: '%s' (%d bytes)", file_name, len(file_bytes))
        logger.info("[IMPORT DEBUG] Mapped fields: %r", column_mapping)
        logger.info("[IMPORT DEBUG] Import mode: '%s', strategy: '%s'", mode, strategy)

        # K0 fast-path settings
        effective_header_row = header_row
        effective_sheet = sheet_name
        effective_uk_col: int | None = None
        if is_k0:
            effective_header_row = effective_header_row or K0_HEADER_ROW
            effective_sheet = effective_sheet or K0_SHEET_NAME
            effective_uk_col = K0_UNIQUE_KEY_COL
            # For K0, always use the deterministic positional mapping
            effective_column_mapping: dict[str, str] = {}
        else:
            effective_column_mapping = column_mapping

        preview = await self.preview_and_validate(
            file_bytes=file_bytes,
            file_name=file_name,
            entity_type=entity_type,
            custom_mapping=effective_column_mapping if not is_k0 else None,
            orientation=orientation,
            sheet_name=effective_sheet,
            header_row=effective_header_row,
        )

        validation_errors = preview["validation_errors"]
        has_errors = len(validation_errors) > 0
        logger.info(
            "[IMPORT DEBUG] Validation summary: Total Rows=%d, Valid=%d, Errors=%d",
            preview.get("total_rows", 0),
            preview.get("valid_rows_count", 0),
            len(validation_errors),
        )

        if strategy == "rollback_all" and has_errors:
            logger.warning("[IMPORT DEBUG] Aborting import due to 'rollback_all' strategy with %d validation error(s)", len(validation_errors))
            raise ValueError(
                f"Import aborted: {len(validation_errors)} error(s) detected and 'Rollback All' strategy was selected."
            )

        parsed = self.parse_excel_file(
            file_bytes,
            specified_orientation=orientation,
            specified_sheet_name=effective_sheet,
            specified_header_row=effective_header_row,
            unique_key_col_index=effective_uk_col,
            max_source_columns=K0_SOURCE_COLUMN_COUNT if is_k0 else None,
        )
        data_rows = parsed["rows"]
        headers = parsed["headers"]
        schema = await self._get_schema(entity_type)
        unique_key_spec = next((c for c in schema.columns if c.unique_key), None)

        # Build column_mapping for the execution loop
        if is_k0:
            # Use index-based mapping (immune to duplicate header collisions)
            k0_execute_index_map = _build_k0_index_mapping(headers)
            column_mapping = _build_k0_column_mapping(headers)  # shim for any non-loop callers
            logger.info(
                "[K0 IMPORT] processing_data_rows=%d",
                len(data_rows),
            )
        elif effective_column_mapping:
            column_mapping = self._normalize_column_mapping(effective_column_mapping, schema, headers)
            k0_execute_index_map = {}
        else:
            column_mapping = self.auto_map_columns(headers, schema)
            k0_execute_index_map = {}
        logger.info("[IMPORT DEBUG] Rows detected: %d", len(data_rows))
        logger.info("[IMPORT DEBUG] Headers detected: %r", headers)
        logger.info("[IMPORT DEBUG] Column mapping: %r", column_mapping)

        # Categorize validation errors for row-level skip logic
        is_project_entity = entity_type.lower() not in ("suppliers", "parts", "risks", "capacity")

        excel_dup_map = {
            err["row_index"]: err["message"]
            for err in validation_errors
            if err.get("error_type") == "DuplicateInFile" and err["row_index"] > 0
        }
        db_dup_map = {
            err["row_index"]: err["message"]
            for err in validation_errors
            if err.get("error_type") == "DuplicateInDatabase" and err["row_index"] > 0
        }
        val_error_map = {
            err["row_index"]: err
            for err in validation_errors
            if err.get("error_type") not in ("DuplicateInFile", "DuplicateInDatabase") and err["row_index"] > 0
        }

        imported_count = 0
        updated_count = 0
        skipped_count = 0
        failed_count = 0
        duplicate_in_excel_count = 0
        duplicate_in_db_count = 0
        validation_errors_count = 0
        non_record_rows_count = 0

        skipped_details: list[dict[str, Any]] = []
        serializer_errors: list[dict[str, Any]] = []

        # Runtime batch tracking for inserted unique keys
        inserted_in_batch: set[str] = set()

        from app.application.services.header_normalizer import normalize_header

        norm_column_mapping: dict[str, str] = {}
        for m_hdr, m_k in column_mapping.items():
            if m_hdr and m_k:
                norm_column_mapping[normalize_header(m_hdr)] = m_k

        logger.info("[IMPORT DEBUG] Database transaction: Starting transaction session...")
        async with self._uow as uow:
            for row_idx, row_vals in enumerate(data_rows, start=1):
                # Prepare row values (field key -> raw value).
                # For K0: use index-based mapping so duplicate header strings can't collide.
                # Rows are already bounded to K0_SOURCE_COLUMN_COUNT by parse_excel_file.
                row_dict: dict[str, Any] = {}
                if is_k0:
                    for col_idx, val in enumerate(row_vals):
                        db_key = k0_execute_index_map.get(col_idx)
                        if db_key:
                            row_dict[db_key] = val
                else:
                    for header, val in zip(headers, row_vals):
                        db_key = column_mapping.get(header)
                        if not db_key:
                            norm_h = normalize_header(header)
                            db_key = norm_column_mapping.get(norm_h)
                        if db_key:
                            row_dict[db_key] = val

                # 0. Skip non-record rows (no value in any unique-key column).
                if not self._row_has_record_key(row_dict, schema):
                    skipped_count += 1
                    non_record_rows_count += 1
                    skipped_details.append({
                        "row_index": row_idx,
                        "reason": "Ignored non-record row",
                        "message": "Row has no value for the record's unique key; skipped.",
                    })
                    continue

                # 1. Skip Duplicate in Excel file
                if row_idx in excel_dup_map:
                    skipped_count += 1
                    duplicate_in_excel_count += 1
                    skipped_details.append({
                        "row_index": row_idx,
                        "reason": "Duplicate in Excel",
                        "message": excel_dup_map[row_idx],
                    })
                    continue

                # 2. Skip Duplicate in Database (ONLY for non-project entities when mode == "insert")
                if not is_project_entity and mode == "insert" and row_idx in db_dup_map:
                    skipped_count += 1
                    duplicate_in_db_count += 1
                    skipped_details.append({
                        "row_index": row_idx,
                        "reason": "Already exists in database",
                        "message": db_dup_map[row_idx],
                    })
                    continue

                # 3. Skip General Validation Error
                if row_idx in val_error_map:
                    skipped_count += 1
                    validation_errors_count += 1
                    err_info = val_error_map[row_idx]
                    col_name = err_info.get("column_name", "")
                    skipped_details.append({
                        "row_index": row_idx,
                        "reason": f"Validation error: {col_name}" if col_name else "Validation error",
                        "message": err_info.get("message", "Validation error"),
                        "column_name": col_name,
                        "field_key": err_info.get("field_key"),
                        "raw_value": err_info.get("raw_value"),
                    })
                    continue

                # Coerce data types via DataNormalizer pipeline
                coerced_row: dict[str, Any] = {}
                for col_spec in schema.columns:
                    if col_spec.key in row_dict:
                        raw_val = row_dict[col_spec.key]
                        norm = DataNormalizer.normalize(raw_val, col_spec)
                        if not norm.is_null:
                            c_val, err_msg = self._validate_and_coerce_type(norm.normalized_value, col_spec)
                            if err_msg:
                                serializer_errors.append({
                                    "row_index": row_idx,
                                    "field_key": col_spec.key,
                                    "column_name": col_spec.label,
                                    "raw_value": str(raw_val),
                                    "error": err_msg,
                                })
                            coerced_row[col_spec.key] = c_val
                        else:
                            coerced_row[col_spec.key] = None

                # Also preserve any non-schema dynamic fields present in row_dict
                for k, v in row_dict.items():
                    if k not in coerced_row and v is not None:
                        coerced_row[k] = v

                # 4. Check runtime duplicate in current batch run
                u_val_str = None
                clean_u_val = None
                if unique_key_spec and unique_key_spec.key in coerced_row and coerced_row[unique_key_spec.key] is not None:
                    u_val_str = str(coerced_row[unique_key_spec.key]).strip()
                    clean_u_val = u_val_str.lower()

                if not is_project_entity and mode == "insert" and clean_u_val and clean_u_val in inserted_in_batch:
                    skipped_count += 1
                    duplicate_in_excel_count += 1
                    skipped_details.append({
                        "row_index": row_idx,
                        "reason": "Duplicate in Excel",
                        "message": f"Duplicate value '{u_val_str}' already inserted in current import batch.",
                    })
                    continue

                # 5. Row-level isolated database operation (SAVEPOINT)
                try:
                    async with uow.session.begin_nested():
                        success, is_update = await self._import_single_record(
                            uow=uow,
                            entity_type=entity_type,
                            record=coerced_row,
                            mode=mode,
                            schema=schema,
                        )

                    code_val = self._extract_project_code(coerced_row, schema) if is_project_entity else (u_val_str or str(row_idx))
                    if success:
                        if is_update:
                            updated_count += 1
                        else:
                            imported_count += 1
                            if clean_u_val:
                                inserted_in_batch.add(clean_u_val)
                        logger.info(
                            "[IMPORT DEBUG]\nrow=%d\nproject_code=%s\ndecision=%s\nreason=%s\nexisting_project_id=%s\nexisting_project_code=%s\ndeleted/active=%s\nvalidation_errors=[]\nmapped_data=%r",
                            row_idx,
                            code_val,
                            "UPDATED" if is_update else "IMPORTED",
                            "Record updated/restored in database" if is_update else "New record created in database",
                            "RESOLVED" if is_update else "None",
                            code_val if is_update else "None",
                            "RESOLVED" if is_update else "NONE",
                            coerced_row,
                        )
                    else:
                        skipped_count += 1
                        reason_msg = "Missing unique identifier" if is_project_entity else "Already exists in database"
                        detail_msg = f"Row {row_idx} is missing a valid project identifier." if is_project_entity else f"Record '{u_val_str or row_idx}' already exists in database."
                        skipped_details.append({
                            "row_index": row_idx,
                            "reason": reason_msg,
                            "message": detail_msg,
                        })
                        if not is_project_entity:
                            duplicate_in_db_count += 1
                        logger.warning(
                            "[IMPORT DEBUG]\nrow=%d\nproject_code=%s\ndecision=SKIPPED\nreason=%s\nexisting_project_id=None\nexisting_project_code=None\ndeleted/active=NONE\nvalidation_errors=[]\nmapped_data=%r",
                            row_idx,
                            code_val or "None",
                            reason_msg,
                            coerced_row,
                        )

                except Exception as e:
                    # SAVEPOINT automatically rolled back only this row's savepoint
                    failed_count += 1
                    ser_err = {
                        "row_index": row_idx,
                        "error": str(e),
                        "record": self._make_json_safe(coerced_row),
                    }
                    serializer_errors.append(ser_err)
                    logger.error("[IMPORT DEBUG] Row %d isolated DB insertion error: %s", row_idx, str(e), exc_info=True)
                    if strategy == "rollback_all":
                        logger.warning("[IMPORT DEBUG] Database transaction: Rolling back transaction due to row %d error", row_idx)
                        raise e

            duration_ms = int((time.time() - start_time) * 1000)
            logger.info("[IMPORT DEBUG] Serializer errors count: %d", len(serializer_errors))

            # Record Audit Log
            await uow.activity_logs.create(
                {
                    "user_id": user_id,
                    "action": ActivityAction.CREATE.value,
                    "resource_type": f"import_{entity_type}",
                    "resource_id": str(uuid.uuid4()),
                    "details": {
                        "file_name": file_name,
                        "entity_type": entity_type,
                        "imported_count": imported_count,
                        "updated_count": updated_count,
                        "skipped_count": skipped_count,
                        "failed_count": failed_count,
                        "non_record_rows_count": non_record_rows_count,
                        "total_rows": len(data_rows),
                        "user_email": user_email,
                    },
                }
            )

            # Record Import History
            history_entry = await uow.import_history.create(
                {
                    "user_id": user_id,
                    "user_email": user_email,
                    "entity_type": entity_type,
                    "file_name": file_name,
                    "file_size": len(file_bytes),
                    "total_rows": len(data_rows),
                    "imported_count": imported_count,
                    "updated_count": updated_count,
                    "skipped_count": skipped_count,
                    "failed_count": failed_count,
                    "duration_ms": duration_ms,
                    "mode": mode,
                    "strategy": strategy,
                    "status": "completed" if failed_count == 0 else "partial",
                    "errors_summary": validation_errors[:20],
                }
            )

            # Dispatch Notification to User
            if user_id:
                try:
                    await uow.notifications.create(
                        {
                            "user_id": user_id,
                            "title": f"Import Completed: {entity_type}",
                            "message": f"Processed '{file_name}': {imported_count} imported, {updated_count} updated, {skipped_count} skipped.",
                            "type": "success" if failed_count == 0 else "warning",
                            "link": "/projects",
                            "is_read": False,
                        }
                    )
                except Exception as notif_err:
                    logger.warning("Failed to create notification on import: %s", notif_err)

            await uow.commit()
            logger.info(
                "[IMPORT DEBUG] Database transaction: Committed successfully (Imported: %d, Updated: %d, Skipped: %d, Failed: %d)",
                imported_count,
                updated_count,
                skipped_count,
                failed_count,
            )

        logger.info("[IMPORT DEBUG] Execution time: %d ms", duration_ms)
        if is_k0:
            logger.info(
                "[K0 IMPORT] imported=%d updated=%d skipped=%d failed=%d",
                imported_count, updated_count, skipped_count, failed_count,
            )
            for skip in skipped_details:
                logger.info(
                    "[K0 IMPORT] row=%d skipped reason='%s' detail='%s'",
                    skip.get("row_index", 0),
                    skip.get("reason", ""),
                    skip.get("message", ""),
                )

        return {
            "id": str(history_entry.id),
            "entity_type": entity_type,
            "file_name": file_name,
            "total_rows": len(data_rows),
            "physical_rows": preview.get("physical_rows", len(data_rows)),
            "empty_rows_count": preview.get("empty_rows_count", 0),
            "non_record_rows_count": non_record_rows_count,
            "record_rows_count": preview.get("record_rows_count", 0),
            "imported_count": imported_count,
            "updated_count": updated_count,
            "skipped_count": skipped_count,
            "failed_count": failed_count,
            "duplicate_in_excel_count": duplicate_in_excel_count,
            "duplicate_in_db_count": duplicate_in_db_count,
            "validation_errors_count": validation_errors_count,
            "skipped_details": skipped_details[:50],
            "duration_ms": duration_ms,
            "status": "completed" if failed_count == 0 else "partial",
            "serializer_errors": serializer_errors,
            "message": f"Import completed: {imported_count} imported, {updated_count} updated, {skipped_count} skipped, {failed_count} failed.",
        }

    def generate_sample_template(self, entity_type: str) -> bytes:
        """
        Creates a clean Excel template formatted with headers, example rows, and instructions.
        """
        if entity_type not in ENTITY_IMPORT_SCHEMAS:
            raise ValueError(f"Unknown entity type: '{entity_type}'")

        schema = ENTITY_IMPORT_SCHEMAS[entity_type]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{schema.display_name} Import"

        # Styling
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        req_font = Font(name="Calibri", size=11, bold=True, color="FFD700") # Gold for required
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        # Write Headers
        headers = [c.label + (" *" if c.required else "") for c in schema.columns]
        ws.append(headers)

        for col_idx, col_spec in enumerate(schema.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = req_font if col_spec.required else header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Write Sample Rows
        for sample in schema.sample_rows:
            row_vals = [sample.get(c.label, "") for c in schema.columns]
            ws.append(row_vals)

        # Autofit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def _get_existing_db_records(
        self, entity_type: str
    ) -> dict[str, dict[str, Any]]:
        """
        Retrieves existing database records indexed by normalized unique key,
        INCLUDING soft-deleted records for accurate CREATE, UPDATE, and RESTORE operations.
        For project entities, scopes results to the matching template code so that
        K9 imports never collide with K0 records and vice versa.
        Returns: { clean_key: {"id": obj_id, "is_deleted": bool, "code": key_str} }
        """
        records: dict[str, dict[str, Any]] = {}

        if entity_type == "suppliers":
            from app.infrastructure.persistence.models.supplier import Supplier
            from sqlalchemy import select
            stmt = select(Supplier)
            res = await self._uow.session.execute(stmt)
            for s in res.scalars().all():
                if s.name:
                    clean = s.name.strip().lower()
                    is_del = bool(getattr(s, "deleted_at", None) is not None)
                    records[clean] = {"id": s.id, "is_deleted": is_del, "name": s.name}

        elif entity_type == "parts":
            from app.infrastructure.persistence.models.project_part import ProjectPart
            from sqlalchemy import select
            stmt = select(ProjectPart)
            res = await self._uow.session.execute(stmt)
            for pt in res.scalars().all():
                if pt.part_number:
                    clean = pt.part_number.strip().lower()
                    records[clean] = {
                        "id": pt.id,
                        "is_deleted": pt.deleted_at is not None,
                        "part_number": pt.part_number,
                        "project_id": pt.project_id,
                    }

        elif entity_type == "risks":
            from app.infrastructure.persistence.models.risk import Risk
            from sqlalchemy import select
            stmt = select(Risk)
            res = await self._uow.session.execute(stmt)
            for r in res.scalars().all():
                if r.title:
                    clean = r.title.strip().lower()
                    is_del = bool(getattr(r, "deleted_at", None) is not None)
                    records[clean] = {"id": r.id, "is_deleted": is_del, "title": r.title}

        else:
            # Query projects scoped to the importing template code
            # so K9 rows never match K0 projects and vice versa
            tmpl_code = entity_type.upper()
            from app.infrastructure.persistence.models.project import Project as ProjectModel
            from sqlalchemy import select
            stmt = select(ProjectModel)
            res = await self._uow.session.execute(stmt)
            for p in res.scalars().all():
                # Skip projects that belong to a DIFFERENT template
                p_tmpl = None
                if p.data and isinstance(p.data, dict):
                    p_tmpl = str(p.data.get("template_code") or "").upper().strip()
                if p_tmpl and p_tmpl != tmpl_code:
                    continue  # cross-template — never match

                is_del = p.deleted_at is not None
                info = {"id": p.id, "is_deleted": is_del, "code": p.code}
                if p.code:
                    records[p.code.strip().lower()] = info
                if p.data and isinstance(p.data, dict):
                    for k in ("unique_id", "code", "part_number"):
                        v = p.data.get(k)
                        if v:
                            records[str(v).strip().lower()] = info

        return records


    async def _get_existing_db_unique_keys(
        self, entity_type: str, unique_key: str
    ) -> set[str]:
        records = await self._get_existing_db_records(entity_type)
        return set(records.keys())

    async def _import_single_record(
        self,
        uow: UnitOfWork,
        entity_type: str,
        record: dict[str, Any],
        mode: str,
        schema: EntityImportSchema | None = None,
    ) -> tuple[bool, bool]:
        """
        Inserts, updates, or restores a record depending on mode and existing presence.
        Returns tuple: (success: bool, is_update: bool)
        """
        if entity_type == "suppliers":
            name = record.get("name")
            existing = await uow.suppliers.get_by_name(name) if name else None
            if existing:
                if mode == "upsert":
                    await uow.suppliers.update(existing.id, record)
                    return True, True
                else:
                    return False, False
            else:
                await uow.suppliers.create(record)
                return True, False

        elif entity_type == "parts":
            part_number = record.get("part_number")
            project_code = record.pop("project_code", None)
            
            project_id = None
            if project_code:
                project = await uow.projects.get_by_code(project_code)
                if project:
                    project_id = project.id
                    record["project_id"] = project_id
                else:
                    return False, False  # Skip if invalid project code
            else:
                first_project = (await uow.projects.get_multi(limit=1))
                if first_project:
                    project_id = first_project[0].id
                    record["project_id"] = project_id
                else:
                    return False, False

            existing = await uow.project_parts.get_by_part_number(part_number) if part_number else None
            if existing and existing.project_id != project_id:
                existing = None

            if existing:
                if mode == "upsert":
                    await uow.project_parts.update(existing.id, record)
                    return True, True
                else:
                    return False, False
            else:
                await uow.project_parts.create(record)
                return True, False

        elif entity_type == "risks":
            title = record.get("title")
            risks = await uow.risks.get_multi(filters={"title": title}) if title else []
            existing = risks[0] if risks else None
            if existing:
                if mode == "upsert":
                    await uow.risks.update(existing.id, record)
                    return True, True
                else:
                    return False, False
            else:
                await uow.risks.create(record)
                return True, False

        else:
            # Default to Project / Project Template (CREATE + UPDATE + RESTORE)
            code = self._extract_project_code(record, schema)
            if not code or str(code).strip() == "":
                # Missing unique project identifier -> fail this record with explicit reason
                logger.warning(
                    "[IMPORT DEBUG] Project code resolution failed for record: %r",
                    record
                )
                return False, False

            name = self._extract_project_name(record, code)
            description = self._extract_project_description(record)
            client_name = self._extract_client_name(record)

            tmpl = await uow.templates.get_by_code(entity_type.upper())
            if tmpl is None:
                try:
                    tmpl = await uow.templates.get(uuid.UUID(entity_type.strip()))
                except (ValueError, AttributeError):
                    pass

            tmpl_code = tmpl.code if tmpl else entity_type.upper()
            inner_data = self._make_json_safe(record)
            if not isinstance(inner_data, dict):
                inner_data = {}
            inner_data["template_code"] = tmpl_code

            from app.application.services.project_service import (
                calculate_workflow_step,
                K0_BUYER_FIELDS,
                K0_CAPACITY_FIELDS,
                K0_SQD_FIELDS,
                _has_any_field,
            )
            workflow_step = calculate_workflow_step(inner_data, tmpl_code)
            inner_data["workflow_step"] = workflow_step

            project_payload: dict[str, Any] = {
                "code": str(code).strip(),
                "name": str(name).strip(),
                "description": description,
                "client_name": client_name,
                "notes": record.get("notes"),
                "data": inner_data,
            }
            if tmpl:
                project_payload["template_id"] = tmpl.id
                project_payload["template_version"] = tmpl.version

            code_str = str(code).strip()
            from sqlalchemy import select, func
            from app.infrastructure.persistence.models.project import Project as ProjectModel

            # Look up project in database INCLUDING soft-deleted records (case-insensitive)
            clean_target = code_str.lower()
            stmt = select(ProjectModel).where(func.lower(ProjectModel.code) == clean_target)
            res = await uow.session.execute(stmt)
            existing_obj = res.scalars().first()

            has_buyer = _has_any_field(inner_data, K0_BUYER_FIELDS)
            has_capacity = _has_any_field(inner_data, K0_CAPACITY_FIELDS)
            has_sqd = _has_any_field(inner_data, K0_SQD_FIELDS)

            if existing_obj:
                # Existing or soft-deleted project found: UPDATE + RESTORE
                is_soft_del = existing_obj.deleted_at is not None
                decision = "RESTORED" if is_soft_del else "UPDATED"
                reason = "Soft-deleted project restored and updated with Excel values" if is_soft_del else "Active project updated with Excel values"

                from sqlalchemy.orm.attributes import flag_modified
                from datetime import datetime, timezone

                for field, value in project_payload.items():
                    if hasattr(existing_obj, field) and value is not None:
                        setattr(existing_obj, field, value)

                if hasattr(existing_obj, "data") and "data" in project_payload:
                    merged_data = dict(existing_obj.data or {})
                    merged_data.update(project_payload["data"] or {})
                    merged_data["template_code"] = tmpl_code
                    merged_data["workflow_step"] = calculate_workflow_step(merged_data, tmpl_code)
                    existing_obj.data = merged_data
                    flag_modified(existing_obj, "data")
                    # Auto-set project status based on workflow completion
                    if merged_data["workflow_step"] == 4:
                        existing_obj.status = "completed"
                    elif existing_obj.status in (None, "draft"):
                        existing_obj.status = "active"

                # Explicitly restore soft-deleted record and update timestamp
                existing_obj.deleted_at = None
                existing_obj.updated_at = datetime.now(timezone.utc)

                uow.session.add(existing_obj)
                await uow.session.flush()

                logger.info(
                    "[IMPORT DEBUG] Project execution decision:\nproject_code=%s\ndecision=%s\nreason=%s\nexisting_project_id=%s\nexisting_project_code=%s\ndeleted/active=%s\nmapped_data=%r",
                    code_str,
                    decision,
                    reason,
                    str(existing_obj.id),
                    existing_obj.code,
                    "DELETED" if is_soft_del else "ACTIVE",
                    record,
                )
                logger.info(
                    "[K0 IMPORT WORKFLOW]\nproject_code=%s\nproject_id=%s\nstructure_id=%s\nstructure_code=%s\n\nbuyer_module=Buyer\nbuyer_data_created=%s\n\ncapacity_module=Capacity Manager\ncapacity_data_created=%s\n\nsqd_module=SQD\nsqd_data_created=%s\n\ncurrent_step=%s\nworkflow_state=%s",
                    code_str,
                    str(existing_obj.id),
                    str(tmpl.id) if tmpl else "None",
                    tmpl_code,
                    "true" if has_buyer else "false",
                    "true" if has_capacity else "false",
                    "true" if has_sqd else "false",
                    workflow_step,
                    f"Step {workflow_step}",
                )
                return True, True  # success=True, is_update=True
            else:
                # New project: CREATE
                # Auto-set status based on workflow step
                if workflow_step == 4:
                    project_payload["status"] = "completed"
                else:
                    project_payload["status"] = "active"
                new_proj = await uow.projects.create(project_payload)
                logger.info(
                    "[IMPORT DEBUG] Project execution decision:\nproject_code=%s\ndecision=CREATED\nreason=New project created\nexisting_project_id=None\nexisting_project_code=None\ndeleted/active=NONE\nmapped_data=%r",
                    code_str,
                    record,
                )
                logger.info(
                    "[K0 IMPORT WORKFLOW]\nproject_code=%s\nproject_id=%s\nstructure_id=%s\nstructure_code=%s\n\nbuyer_module=Buyer\nbuyer_data_created=%s\n\ncapacity_module=Capacity Manager\ncapacity_data_created=%s\n\nsqd_module=SQD\nsqd_data_created=%s\n\ncurrent_step=%s\nworkflow_state=%s",
                    code_str,
                    str(new_proj.id),
                    str(tmpl.id) if tmpl else "None",
                    tmpl_code,
                    "true" if has_buyer else "false",
                    "true" if has_capacity else "false",
                    "true" if has_sqd else "false",
                    workflow_step,
                    f"Step {workflow_step}",
                )
                return True, False  # success=True, is_update=False

    def _validate_and_coerce_type(
        self, val: Any, spec: ImportColumnSpec
    ) -> tuple[Any, str | None]:
        if val is None:
            return None, None

        field_type = (spec.type or "string").lower()

        if field_type in ("string", "text", "textarea"):
            val_str = val.isoformat() if isinstance(val, (datetime, date)) else str(val).strip()
            return val_str, None

        if isinstance(val, (datetime, date)):
            d_val = val.date() if isinstance(val, datetime) else val
            return d_val, None

        val_str = str(val).strip()
        if not val_str:
            return None, None

        elif field_type == "integer":
            if isinstance(val, int) and not isinstance(val, bool):
                return val, None
            try:
                f_val = float(val_str)
                return int(f_val), None
            except (ValueError, OverflowError):
                return None, f"Expected integer number for '{spec.label}', got '{val_str}'"

        elif field_type == "number":
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                return float(val), None
            try:
                return float(val_str), None
            except (ValueError, OverflowError):
                return None, f"Expected numeric value for '{spec.label}', got '{val_str}'"

        elif field_type == "enum":
            clean_val = val_str.strip()
            if spec.enum_values:
                clean_clean = clean_val.lower().replace(" ", "_")
                matched_enum = next((ev for ev in spec.enum_values if ev.lower().replace(" ", "_") == clean_clean), None)
                if matched_enum is not None:
                    return matched_enum, None
                allowed = ", ".join(spec.enum_values)
                return None, f"Invalid value '{val_str}'. Allowed values: {allowed}"
        elif field_type == "week" or spec.key in ("cat1_forecast_date_cw", "cat2_forecast_date", "cat3_forecast_date"):
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                return int(val), None
            if isinstance(val, (datetime, date)):
                d_val = val.date() if isinstance(val, datetime) else val
                return d_val.isoformat(), None
            clean_s = str(val).strip()
            if re.match(r"^\d+$", clean_s):
                return int(clean_s), None
            if clean_s:
                return clean_s, None
            return None, None

        elif field_type == "date":
            # For date types, if DataNormalizer returned a date, it was already handled above.
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                return int(val), None
            if isinstance(val, str) and re.match(r"^\d+$", val.strip()):
                return int(val.strip()), None
            # If it's a string here, it failed date normalization.
            return None, f"Invalid date format for '{spec.label}'. Expected YYYY-MM-DD."

        return val_str, None
