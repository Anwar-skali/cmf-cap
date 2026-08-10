"""
test_vertical_import.py
=======================
Tests for vertical / key-value Excel orientation detection and extraction.

The tests exercise:
  1. ExcelHeaderExtractor.detect_orientation() with a synthetic vertical sheet
  2. ExcelHeaderExtractor.detect_orientation() with a standard horizontal sheet
  3. ExcelHeaderExtractor.detect_orientation() with user-specified override
  4. ExcelHeaderExtractor.extract_headers_with_details() in VERTICAL mode
     — verifies Col A values are used as headers, Col B values are NOT headers
  5. ImportEngineService.parse_excel_file() produces a single-row normalised
     record for vertical files and multi-row for horizontal files

All tests use in-memory openpyxl workbooks — no real files are read from disk.
"""
import io
import types
import pytest
import openpyxl

from app.application.services.excel_header_extractor import ExcelHeaderExtractor
from app.application.services.import_engine_service import ImportEngineService


# ─── Helper: build in-memory Excel workbooks ─────────────────────────────────

def _build_vertical_wb() -> bytes:
    """
    K9-style vertical workbook:
      Column A = field name  |  Column B = value
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "K9 Data"
    rows = [
        ("Unique_ID",        "UID-V-0001"),
        ("Part_Name",        "Brake Module"),
        ("Part_Number",      "PN-2024-BM"),
        ("Supplier_Name",    "ACME Suppliers"),
        ("APQP",             "Completed"),
        ("Weekly_Capacity",  1500),
        ("Status",           "Active"),
        ("Date",             "2024-01-15"),
        ("Category",         "Safety"),
        ("Description",      "Front left brake assembly"),
        ("Version",          "2.1"),
        ("Code",             "BM-001"),
    ]
    for field, value in rows:
        ws.append([field, value])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_horizontal_wb() -> bytes:
    """
    Standard horizontal workbook:
      Row 1 = headers | Rows 2+ = data records
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"
    ws.append(["Unique_ID", "Part_Name", "Supplier_Name", "Status", "Weekly_Capacity"])
    ws.append(["UID-H-001", "Door Panel",  "SupA", "Active",  2000])
    ws.append(["UID-H-002", "Seat Frame",  "SupB", "Active",  1800])
    ws.append(["UID-H-003", "Roof Module", "SupC", "Pending",  900])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rows_from_bytes(wb_bytes: bytes, sheet_name: str | None = None) -> list[list]:
    """Load workbook and return rows as list[list]."""
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


# ─── detect_orientation tests ─────────────────────────────────────────────────

class TestDetectOrientation:

    def test_vertical_detected(self):
        rows = _rows_from_bytes(_build_vertical_wb())
        result = ExcelHeaderExtractor.detect_orientation(rows)
        assert result["orientation"] == "VERTICAL", (
            f"Expected VERTICAL but got {result['orientation']}. Reason: {result.get('reason')}"
        )
        assert result["orientation_confidence"] >= 50

    def test_horizontal_detected(self):
        rows = _rows_from_bytes(_build_horizontal_wb())
        result = ExcelHeaderExtractor.detect_orientation(rows)
        assert result["orientation"] == "HORIZONTAL", (
            f"Expected HORIZONTAL but got {result['orientation']}. Reason: {result.get('reason')}"
        )

    def test_user_override_forces_vertical(self):
        """Even a horizontal-looking file must report VERTICAL when user specifies."""
        rows = _rows_from_bytes(_build_horizontal_wb())
        result = ExcelHeaderExtractor.detect_orientation(rows, specified_orientation="VERTICAL")
        assert result["orientation"] == "VERTICAL"
        assert result["orientation_confidence"] == 100

    def test_user_override_forces_horizontal(self):
        """Even a vertical-looking file must report HORIZONTAL when user specifies."""
        rows = _rows_from_bytes(_build_vertical_wb())
        result = ExcelHeaderExtractor.detect_orientation(rows, specified_orientation="HORIZONTAL")
        assert result["orientation"] == "HORIZONTAL"
        assert result["orientation_confidence"] == 100


# ─── extract_headers_with_details tests ──────────────────────────────────────

class TestExtractHeadersVertical:
    """
    extract_headers_with_details returns an 8-tuple:
      (headers, detected_row_idx, header_confidence, sheet_used,
       sheet_confidence, row_previews, duration_ms, orientation_info)
    """

    def test_vertical_extraction_returns_field_names(self):
        headers, *_, orientation_info = ExcelHeaderExtractor.extract_headers_with_details(
            _build_vertical_wb(),
            specified_orientation="VERTICAL",
        )
        assert "Unique_ID" in headers
        assert "Part_Name" in headers
        assert "Supplier_Name" in headers
        assert "APQP" in headers

    def test_vertical_extraction_no_col_b_values_in_headers(self):
        """Column B values must NEVER appear as header names."""
        headers, *_ = ExcelHeaderExtractor.extract_headers_with_details(
            _build_vertical_wb(),
            specified_orientation="VERTICAL",
        )
        assert "UID-V-0001" not in headers, "Column B value leaked into headers"
        assert "Brake Module" not in headers, "Column B value leaked into headers"
        assert "ACME Suppliers" not in headers, "Column B value leaked into headers"

    def test_vertical_extraction_no_empty_headers(self):
        headers, *_ = ExcelHeaderExtractor.extract_headers_with_details(
            _build_vertical_wb(),
            specified_orientation="VERTICAL",
        )
        for h in headers:
            assert h.strip() != "", f"Empty header found: {h!r}"

    def test_vertical_orientation_info_returned(self):
        *_, orientation_info = ExcelHeaderExtractor.extract_headers_with_details(
            _build_vertical_wb(),
            specified_orientation="VERTICAL",
        )
        assert orientation_info["orientation"] == "VERTICAL"


# ─── ImportEngineService.parse_excel_file tests ───────────────────────────────

def _make_engine():
    """Create an ImportEngineService with a stub UoW (parse_excel_file doesn't need DB)."""
    stub_uow = types.SimpleNamespace()
    return ImportEngineService(stub_uow)


class TestParseExcelFileVertical:

    def test_vertical_parse_produces_single_row(self):
        engine = _make_engine()
        parsed = engine.parse_excel_file(
            file_bytes=_build_vertical_wb(),
            specified_orientation="VERTICAL",
        )
        headers = parsed["headers"]
        data_rows = parsed["rows"]

        assert len(data_rows) == 1, (
            f"Vertical import must produce exactly 1 data row, got {len(data_rows)}"
        )
        assert len(headers) > 0
        assert len(data_rows[0]) == len(headers), (
            "Single row value count must match header count"
        )

    def test_vertical_parse_values_intact(self):
        engine = _make_engine()
        parsed = engine.parse_excel_file(
            file_bytes=_build_vertical_wb(),
            specified_orientation="VERTICAL",
        )
        row_dict = dict(zip(parsed["headers"], parsed["rows"][0]))
        assert row_dict.get("Unique_ID") == "UID-V-0001"
        assert row_dict.get("Part_Name") == "Brake Module"
        assert row_dict.get("Supplier_Name") == "ACME Suppliers"

    def test_vertical_parse_orientation_metadata(self):
        engine = _make_engine()
        parsed = engine.parse_excel_file(
            file_bytes=_build_vertical_wb(),
            specified_orientation="VERTICAL",
        )
        assert parsed.get("orientation") == "VERTICAL"

    def test_horizontal_parse_unaffected(self):
        """Existing horizontal import must still produce multiple rows."""
        engine = _make_engine()
        parsed = engine.parse_excel_file(
            file_bytes=_build_horizontal_wb(),
            specified_orientation="HORIZONTAL",
        )
        assert len(parsed["rows"]) == 3, (
            f"Horizontal import should produce 3 data rows, got {len(parsed['rows'])}"
        )
        assert parsed.get("orientation") == "HORIZONTAL"

    def test_user_override_reason_text(self):
        """User override must produce 'User-specified Vertical.' or 'User-specified Horizontal.'"""
        r_vert = ExcelHeaderExtractor.detect_orientation([], specified_orientation="VERTICAL")
        assert r_vert["reason"] == "User-specified Vertical."
        assert r_vert["orientation"] == "VERTICAL"
        assert r_vert["orientation_confidence"] == 100.0

        r_horiz = ExcelHeaderExtractor.detect_orientation([], specified_orientation="HORIZONTAL")
        assert r_horiz["reason"] == "User-specified Horizontal."
        assert r_horiz["orientation"] == "HORIZONTAL"
        assert r_horiz["orientation_confidence"] == 100.0

    def test_score_sheet_vertical_orientation_aware(self):
        """Scoring a vertical sheet with specified_orientation='VERTICAL' produces high confidence and classification PROJECT_DATA."""
        wb_bytes = _build_vertical_wb()
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
        ws = wb.active
        info = ExcelHeaderExtractor._score_sheet(ws, ws.title, specified_orientation="VERTICAL")
        assert info["orientation"] == "VERTICAL"
        assert info["confidence"] >= 70.0
        assert info["classification"] == "PROJECT_DATA"
        assert info["project_field_matches"] > 0

